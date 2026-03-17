from __future__ import annotations

import csv
import json
import re
import secrets
import shutil
import sqlite3
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .config import AppPaths
from .database import VoteFreeDB
from .security import hash_passcode, hash_secret, verify_passcode, verify_secret
from .survey_engine import ROSTER_REPEAT_TOKEN, normalize_schema, validate_answers
from .vote_crypto import VoteCrypto, VoteCryptoError


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_iso(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


class ServiceError(Exception):
    pass


@dataclass
class SubmissionResult:
    submission_id: str
    vote_path: Path


class VoteFreeService:
    def __init__(self, paths: AppPaths):
        self.paths = paths
        self.db = VoteFreeDB(paths.db_file)
        self.crypto = VoteCrypto(paths.keys_dir)

    def initialize(self) -> None:
        self.paths.ensure()
        self.db.init_schema()
        if not self.db.get_setting("flask_secret"):
            self.db.set_setting("flask_secret", secrets.token_urlsafe(32))

    def is_bootstrapped(self) -> bool:
        return bool(self.db.get_setting("admin_password_hash")) and self.crypto.keys_exist()

    def initialize_admin(self, password: str) -> None:
        if self.is_bootstrapped():
            raise ServiceError("系统已初始化。")
        self.db.set_setting("admin_password_hash", hash_secret(password))
        self.crypto.generate_keys(password)
        self.db.set_setting("schema_version", "2")
        self.db.append_audit_log("admin_init", {"at": utc_now()})

    def unlock_admin(self, password: str) -> None:
        encoded = self.db.get_setting("admin_password_hash")
        if not encoded or not verify_secret(password, encoded):
            raise ServiceError("管理员密码错误。")
        try:
            self.crypto.unlock_private_key(password)
            self.db.append_audit_log("admin_unlock", {"at": utc_now()})
        except VoteCryptoError as exc:
            raise ServiceError(str(exc)) from exc

    def change_admin_password(self, old_password: str, new_password: str) -> None:
        encoded = self.db.get_setting("admin_password_hash")
        if not encoded or not verify_secret(old_password, encoded):
            raise ServiceError("旧密码错误。")
        if not self.crypto.unlocked:
            self.unlock_admin(old_password)
        self.crypto.reprotect_private_key(new_password)
        self.db.set_setting("admin_password_hash", hash_secret(new_password))
        self.db.append_audit_log("admin_password_changed", {"at": utc_now()})

    def flask_secret(self) -> str:
        value = self.db.get_setting("flask_secret")
        if not value:
            value = secrets.token_urlsafe(32)
            self.db.set_setting("flask_secret", value)
        return value

    def get_runtime_kernel(self) -> str:
        raw = str(self.db.get_setting("runtime_kernel") or "").strip().lower()
        if raw in {"web", "tkinter"}:
            return raw
        return "web"

    def set_runtime_kernel(self, kernel: str) -> str:
        value = str(kernel or "").strip().lower()
        if value not in {"web", "tkinter"}:
            raise ServiceError("内核类型仅支持 web 或 tkinter。")
        self.db.set_setting("runtime_kernel", value)
        self.db.append_audit_log("runtime_kernel_set", {"kernel": value, "at": utc_now()})
        return value

    def toggle_runtime_kernel(self, current_kernel: str) -> str:
        current = str(current_kernel or "").strip().lower()
        if current not in {"web", "tkinter"}:
            current = self.get_runtime_kernel()
        target = "tkinter" if current == "web" else "web"
        return self.set_runtime_kernel(target)

    def _normalize_auth_mode(self, auth_mode: str) -> str:
        mode = auth_mode.strip().lower()
        if mode in {"open", "roster_code", "roster_name_code", "roster_fields"}:
            return mode
        return "open"

    def _normalize_identity_mode(self, identity_mode: str) -> str:
        mode = str(identity_mode or "").strip().lower()
        if mode in {"realname", "semi", "anonymous"}:
            return "realname"
        return "realname"

    def _normalize_collect_field_key(self, raw_key: str, index: int) -> str:
        key = str(raw_key or "").strip()
        if not key:
            return f"field_{index}"
        key = key.replace(" ", "_").replace("-", "_")
        return key

    def _normalize_collect_fields(self, raw_fields: Any) -> List[Dict[str, str]]:
        if not isinstance(raw_fields, list):
            return []
        result: List[Dict[str, str]] = []
        seen: set[str] = set()
        for idx, item in enumerate(raw_fields, start=1):
            if isinstance(item, dict):
                key = self._normalize_collect_field_key(str(item.get("key", "")), idx)
                label = str(item.get("label", "")).strip() or key
            else:
                label = str(item or "").strip()
                key = self._normalize_collect_field_key(label, idx)
            if not key or key in seen:
                continue
            seen.add(key)
            result.append({"key": key, "label": label})
        return result

    def _identity_dedupe_key(self, identity_data: Optional[Dict[str, Any]]) -> str:
        data = identity_data if isinstance(identity_data, dict) else {}
        pairs: List[str] = []
        for key in sorted(data.keys()):
            k = str(key).strip()
            if not k:
                continue
            v = str(data.get(key, "")).strip()
            if not v:
                continue
            pairs.append(f"{k}={v}")
        if not pairs:
            return ""
        return "idv:" + "|".join(pairs)

    def _legacy_collect_fields(self, identity_fields: Dict[str, Any]) -> List[Dict[str, str]]:
        fields: List[Dict[str, str]] = []
        if bool(identity_fields.get("collect_name", False)):
            fields.append({"key": "member_name", "label": "姓名"})
        if bool(identity_fields.get("collect_code", False)):
            fields.append({"key": "member_code", "label": "编号"})
        return fields

    def _normalize_identity_fields(
        self,
        identity_mode: str,
        identity_fields: Optional[Dict[str, Any]],
        auth_mode: str = "open",
        auth_roster_id: str = "",
    ) -> Dict[str, Any]:
        normalized_mode = self._normalize_identity_mode(identity_mode)
        raw = identity_fields if isinstance(identity_fields, dict) else {}
        collect_fields = self._normalize_collect_fields(raw.get("collect_fields"))
        if not collect_fields:
            collect_fields = self._legacy_collect_fields(raw)

        # 兼容旧配置：名单校验模式下若未显式配置采集字段，按校验方式自动补齐。
        if not collect_fields:
            if auth_mode == "roster_name_code":
                collect_fields = [{"key": "member_code", "label": "编号"}, {"key": "member_name", "label": "姓名"}]
            elif auth_mode == "roster_code":
                collect_fields = [{"key": "member_code", "label": "编号"}]
            elif auth_mode == "roster_fields":
                roster_columns = self.get_roster_columns(auth_roster_id)
                if roster_columns:
                    collect_fields = [{"key": c["key"], "label": c["label"]} for c in roster_columns]

        allow_same_device_repeat = bool(raw.get("allow_same_device_repeat", False))
        collect_name = any(str(item.get("key", "")).strip() == "member_name" for item in collect_fields)
        collect_code = any(str(item.get("key", "")).strip() == "member_code" for item in collect_fields)
        return {
            "collect_fields": collect_fields,
            "collect_name": collect_name,
            "collect_code": collect_code,
            "name_required": collect_name,
            "code_required": collect_code,
            "identity_mode": normalized_mode,
            "allow_same_device_repeat": allow_same_device_repeat,
        }

    def _questionnaire_template_key(self, questionnaire: Optional[Dict[str, Any]]) -> str:
        if not isinstance(questionnaire, dict):
            return ""
        schema = questionnaire.get("schema", {})
        if not isinstance(schema, dict):
            return ""
        meta = schema.get("meta", {})
        if not isinstance(meta, dict):
            return ""
        return str(meta.get("template_key", "")).strip()

    def _sync_template_sql_views_to_questionnaire(
        self,
        questionnaire_id: str,
        template_key: str,
        overwrite: bool = False,
    ) -> int:
        qid = str(questionnaire_id or "").strip()
        tkey = str(template_key or "").strip()
        if not qid or not tkey:
            return 0
        q_views = self.db.list_sql_views(qid)
        existing_names = {str(item.get("name", "")).strip() for item in q_views if str(item.get("name", "")).strip()}
        template_views = self.db.list_template_sql_views(tkey)
        synced = 0
        for item in template_views:
            name = str(item.get("name", "")).strip()
            sql_text = str(item.get("sql_text", "")).strip()
            if not name or not sql_text:
                continue
            if not overwrite and name in existing_names:
                continue
            self.db.save_sql_view(qid, name, sql_text)
            synced += 1
        return synced

    def create_questionnaire(
        self,
        title: str,
        description: str,
        identity_mode: str,
        allow_repeat: bool,
        passcode: str,
        schema: Dict[str, Any],
        questionnaire_id: Optional[str] = None,
        auth_mode: str = "open",
        auth_roster_id: str = "",
        identity_fields: Optional[Dict[str, Any]] = None,
    ) -> str:
        qid = questionnaire_id or f"Q{uuid.uuid4().hex[:10]}"
        normalized = normalize_schema(schema)
        existing_q = self.get_questionnaire(qid) if questionnaire_id else None
        normalized_meta = normalized.get("meta", {})
        if not isinstance(normalized_meta, dict):
            normalized_meta = {}
        if isinstance(existing_q, dict):
            old_schema = existing_q.get("schema", {})
            old_meta = old_schema.get("meta", {}) if isinstance(old_schema, dict) else {}
            if isinstance(old_meta, dict):
                old_template_key = str(old_meta.get("template_key", "")).strip()
                if old_template_key and not str(normalized_meta.get("template_key", "")).strip():
                    for key, value in old_meta.items():
                        if str(key).startswith("template_"):
                            normalized_meta.setdefault(str(key), value)
        normalized["meta"] = normalized_meta
        passcode_hash = hash_passcode(passcode.strip()) if passcode.strip() else ""
        normalized_identity_mode = self._normalize_identity_mode(identity_mode)
        normalized_auth_mode = self._normalize_auth_mode(auth_mode)
        normalized_roster_id = auth_roster_id.strip()
        normalized_identity_fields = self._normalize_identity_fields(
            identity_mode=normalized_identity_mode,
            identity_fields=identity_fields,
            auth_mode=normalized_auth_mode,
            auth_roster_id=normalized_roster_id,
        )
        if normalized_auth_mode != "open" and not auth_roster_id.strip():
            raise ServiceError("名单校验模式必须绑定名单。")
        if normalized_roster_id and not self.db.get_roster(normalized_roster_id):
            raise ServiceError("绑定名单不存在，请重新选择。")
        version = self.db.save_questionnaire(
            questionnaire_id=qid,
            title=title.strip(),
            description=description.strip(),
            identity_mode=normalized_identity_mode,
            allow_repeat=allow_repeat,
            passcode_hash=passcode_hash,
            schema=normalized,
            auth_mode=normalized_auth_mode,
            auth_roster_id=normalized_roster_id,
            identity_fields=normalized_identity_fields,
        )
        self.db.append_audit_log(
            "questionnaire_saved",
            {"questionnaire_id": qid, "version": version, "auth_mode": normalized_auth_mode},
        )
        template_key = self._questionnaire_template_key({"schema": normalized})
        if template_key:
            self._sync_template_sql_views_to_questionnaire(qid, template_key, overwrite=False)
        return qid

    def list_questionnaires(self, active_only: bool = False) -> List[Dict[str, Any]]:
        return self.db.list_questionnaires(active_only=active_only)

    def get_questionnaire(self, questionnaire_id: str) -> Optional[Dict[str, Any]]:
        return self.db.get_questionnaire(questionnaire_id)

    def rename_questionnaire(self, questionnaire_id: str, new_title: str) -> int:
        qid = str(questionnaire_id or "").strip()
        title = str(new_title or "").strip()
        if not qid:
            raise ServiceError("问卷不存在。")
        if not title:
            raise ServiceError("问卷标题不能为空。")
        old = self.get_questionnaire(qid)
        if not old:
            raise ServiceError("问卷不存在。")
        version = self.db.save_questionnaire(
            questionnaire_id=qid,
            title=title,
            description=str(old.get("description", "")).strip(),
            identity_mode=self._normalize_identity_mode(str(old.get("identity_mode", "realname"))),
            allow_repeat=bool(old.get("allow_repeat", False)),
            passcode_hash=str(old.get("passcode_hash", "")).strip(),
            schema=normalize_schema(old.get("schema", {})),
            auth_mode=self._normalize_auth_mode(str(old.get("auth_mode", "open"))),
            auth_roster_id=str(old.get("auth_roster_id", "")).strip(),
            identity_fields=old.get("identity_fields", {}),
            status=str(old.get("status", "active")).strip() or "active",
        )
        self.db.append_audit_log(
            "questionnaire_renamed",
            {"questionnaire_id": qid, "new_title": title, "version": version},
        )
        return version

    def copy_questionnaire(
        self,
        source_questionnaire_id: str,
        new_title: str = "",
        new_questionnaire_id: str = "",
    ) -> str:
        src_id = str(source_questionnaire_id or "").strip()
        if not src_id:
            raise ServiceError("源问卷不存在。")
        old = self.get_questionnaire(src_id)
        if not old:
            raise ServiceError("源问卷不存在。")
        qid = str(new_questionnaire_id or "").strip() or f"Q{uuid.uuid4().hex[:10]}"
        if self.get_questionnaire(qid):
            raise ServiceError("目标问卷ID已存在。")
        title = str(new_title or "").strip() or f"{str(old.get('title', '')).strip()}（副本）"
        self.db.save_questionnaire(
            questionnaire_id=qid,
            title=title,
            description=str(old.get("description", "")).strip(),
            identity_mode=self._normalize_identity_mode(str(old.get("identity_mode", "realname"))),
            allow_repeat=bool(old.get("allow_repeat", False)),
            passcode_hash=str(old.get("passcode_hash", "")).strip(),
            schema=normalize_schema(old.get("schema", {})),
            auth_mode=self._normalize_auth_mode(str(old.get("auth_mode", "open"))),
            auth_roster_id=str(old.get("auth_roster_id", "")).strip(),
            identity_fields=old.get("identity_fields", {}),
            status=str(old.get("status", "active")).strip() or "active",
        )
        for item in self.db.list_sql_views(src_id):
            name = str(item.get("name", "")).strip()
            sql_text = str(item.get("sql_text", "")).strip()
            if not name or not sql_text:
                continue
            self.db.save_sql_view(qid, name, sql_text)
        template_key = self._questionnaire_template_key({"schema": old.get("schema", {})})
        if template_key:
            self._sync_template_sql_views_to_questionnaire(qid, template_key, overwrite=False)
        self.db.append_audit_log(
            "questionnaire_copied",
            {"source_questionnaire_id": src_id, "new_questionnaire_id": qid, "new_title": title},
        )
        return qid

    def delete_questionnaire(self, questionnaire_id: str) -> None:
        qid = str(questionnaire_id or "").strip()
        if not qid:
            raise ServiceError("问卷不存在。")
        old = self.get_questionnaire(qid)
        if not old:
            raise ServiceError("问卷不存在。")
        submissions = self.db.list_submissions(questionnaire_id=qid)
        vote_files = [Path(str(item.get("vote_file", "")).strip()) for item in submissions if str(item.get("vote_file", "")).strip()]
        self.db.delete_questionnaire(qid)
        for vote_file in vote_files:
            try:
                if vote_file.exists():
                    vote_file.unlink()
            except Exception:
                continue
        vote_dir = self.paths.votes_dir / qid
        if vote_dir.exists():
            shutil.rmtree(vote_dir, ignore_errors=True)
        self.db.append_audit_log(
            "questionnaire_deleted",
            {"questionnaire_id": qid, "submission_count": len(submissions)},
        )

    def list_questionnaire_versions(self, questionnaire_id: str) -> List[Dict[str, Any]]:
        return self.db.list_questionnaire_versions(questionnaire_id)

    def verify_questionnaire_passcode(self, questionnaire: Dict[str, Any], passcode: str) -> bool:
        code_hash = questionnaire.get("passcode_hash") or ""
        if not code_hash:
            return True
        return verify_passcode(passcode, code_hash)

    def _normalize_roster_columns(self, columns: Optional[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        items = columns if isinstance(columns, list) else []
        result: List[Dict[str, Any]] = []
        seen: set[str] = set()
        for idx, item in enumerate(items, start=1):
            if not isinstance(item, dict):
                continue
            label = str(item.get("label", "")).strip()
            key_raw = str(item.get("key", "")).strip()
            if not label and not key_raw:
                continue
            key = self._normalize_collect_field_key(key_raw or label, idx)
            if key in seen:
                continue
            seen.add(key)
            result.append(
                {
                    "key": key,
                    "label": label or key,
                    "is_key": bool(item.get("is_key", False)),
                }
            )
        if not result:
            result = [
                {"key": "member_name", "label": "姓名", "is_key": False},
                {"key": "member_code", "label": "编号", "is_key": True},
            ]
        key_count = len([x for x in result if bool(x.get("is_key", False))])
        if key_count == 0:
            code_idx = next((i for i, x in enumerate(result) if str(x.get("key", "")).strip() == "member_code"), 0)
            result[code_idx]["is_key"] = True
        elif key_count > 1:
            first = True
            for item in result:
                if not bool(item.get("is_key", False)):
                    continue
                if first:
                    first = False
                    continue
                item["is_key"] = False
        return result

    def _infer_roster_columns_from_members(self, roster_id: str) -> List[Dict[str, Any]]:
        members = self.db.list_roster_members(roster_id, limit=100000)
        has_name = any(str(m.get("member_name", "")).strip() for m in members)
        has_code = any(str(m.get("member_code", "")).strip() for m in members)
        extra_keys: List[str] = []
        seen: set[str] = set()
        for member in members:
            extra = member.get("extra", {})
            if not isinstance(extra, dict):
                continue
            for key in extra.keys():
                k = str(key).strip()
                if not k or k in seen:
                    continue
                seen.add(k)
                extra_keys.append(k)
        result: List[Dict[str, Any]] = []
        if has_name:
            result.append({"key": "member_name", "label": "姓名", "is_key": False})
        if has_code:
            result.append({"key": "member_code", "label": "编号", "is_key": False})
        if not result:
            result.append({"key": "member_key", "label": "唯一标识", "is_key": True})
        for key in extra_keys:
            result.append({"key": key, "label": key, "is_key": False})
        if not any(bool(x.get("is_key", False)) for x in result):
            result[0]["is_key"] = True
        return result

    def get_roster_columns(self, roster_id: str) -> List[Dict[str, Any]]:
        rid = str(roster_id or "").strip()
        if not rid:
            return []
        roster = self.db.get_roster(rid)
        if not roster:
            return []
        normalized = self._normalize_roster_columns(roster.get("columns", []))
        if normalized:
            return normalized
        inferred = self._infer_roster_columns_from_members(rid)
        if inferred:
            self.db.update_roster_columns(rid, json.dumps(inferred, ensure_ascii=False, separators=(",", ":")))
        return inferred

    def set_roster_columns(self, roster_id: str, columns: List[Dict[str, Any]]) -> None:
        rid = str(roster_id or "").strip()
        if not rid:
            raise ServiceError("名单不存在。")
        normalized = self._normalize_roster_columns(columns)
        self.db.update_roster_columns(rid, json.dumps(normalized, ensure_ascii=False, separators=(",", ":")))
        self.db.append_audit_log(
            "roster_columns_updated",
            {"roster_id": rid, "columns": normalized},
        )

    def _member_field_value(self, member: Dict[str, Any], field_key: str) -> str:
        key = str(field_key or "").strip()
        if not key:
            return ""
        if key == "member_name":
            return str(member.get("member_name", "")).strip()
        if key == "member_code":
            return str(member.get("member_code", "")).strip()
        if key == "member_key":
            return str(member.get("member_key", "")).strip()
        extra = member.get("extra", {})
        if isinstance(extra, dict):
            return str(extra.get(key, "")).strip()
        return ""

    def create_roster(
        self,
        name: str,
        description: str = "",
        roster_id: str = "",
        columns: Optional[List[Dict[str, Any]]] = None,
    ) -> str:
        rid = roster_id.strip() or f"R{uuid.uuid4().hex[:10]}"
        if not name.strip():
            raise ServiceError("名单名称不能为空。")
        normalized_columns = self._normalize_roster_columns(columns)
        self.db.create_roster(
            rid,
            name.strip(),
            description.strip(),
            columns_json=json.dumps(normalized_columns, ensure_ascii=False, separators=(",", ":")),
        )
        self.db.append_audit_log("roster_saved", {"roster_id": rid, "name": name.strip()})
        return rid

    def list_rosters(self) -> List[Dict[str, Any]]:
        result = self.db.list_rosters()
        for item in result:
            rid = str(item.get("id", "")).strip()
            if not rid:
                item["columns"] = []
                continue
            item["columns"] = self.get_roster_columns(rid)
        return result

    def rename_roster(self, roster_id: str, new_name: str) -> None:
        rid = str(roster_id or "").strip()
        name = str(new_name or "").strip()
        if not rid:
            raise ServiceError("名单不存在。")
        if not name:
            raise ServiceError("名单名称不能为空。")
        roster = self.db.get_roster(rid)
        if not roster:
            raise ServiceError("名单不存在。")
        self.db.create_roster(
            roster_id=rid,
            name=name,
            description=str(roster.get("description", "")).strip(),
            columns_json=json.dumps(roster.get("columns", []), ensure_ascii=False, separators=(",", ":")),
        )
        self.db.append_audit_log("roster_renamed", {"roster_id": rid, "new_name": name})

    def copy_roster(self, source_roster_id: str, new_name: str = "", new_roster_id: str = "") -> str:
        src_id = str(source_roster_id or "").strip()
        if not src_id:
            raise ServiceError("源名单不存在。")
        source = self.db.get_roster(src_id)
        if not source:
            raise ServiceError("源名单不存在。")
        rid = str(new_roster_id or "").strip() or f"R{uuid.uuid4().hex[:10]}"
        if self.db.get_roster(rid):
            raise ServiceError("目标名单ID已存在。")
        name = str(new_name or "").strip() or f"{str(source.get('name', '')).strip()}（副本）"
        columns = source.get("columns", [])
        self.db.create_roster(
            roster_id=rid,
            name=name,
            description=str(source.get("description", "")).strip(),
            columns_json=json.dumps(columns if isinstance(columns, list) else [], ensure_ascii=False, separators=(",", ":")),
        )
        members = self.db.list_roster_members(src_id, limit=100000)
        if members:
            self.db.upsert_roster_members(roster_id=rid, members=members, replace_all=True)
        self.db.append_audit_log(
            "roster_copied",
            {"source_roster_id": src_id, "new_roster_id": rid, "new_name": name, "member_count": len(members)},
        )
        return rid

    def delete_roster(self, roster_id: str) -> None:
        rid = str(roster_id or "").strip()
        if not rid:
            raise ServiceError("名单不存在。")
        roster = self.db.get_roster(rid)
        if not roster:
            raise ServiceError("名单不存在。")
        refs = [q for q in self.db.list_questionnaires(active_only=False) if str(q.get("auth_roster_id", "")).strip() == rid]
        if refs:
            names = "、".join([str(item.get("title", "")).strip() or str(item.get("id", "")).strip() for item in refs[:3]])
            more = "…" if len(refs) > 3 else ""
            raise ServiceError(f"该名单已被问卷绑定，不能删除：{names}{more}")
        member_count = len(self.db.list_roster_members(rid, limit=100000))
        self.db.delete_roster(rid)
        self.db.append_audit_log(
            "roster_deleted",
            {"roster_id": rid, "member_count": member_count},
        )

    def list_roster_members(self, roster_id: str, limit: int = 5000) -> List[Dict[str, Any]]:
        members = self.db.list_roster_members(roster_id, limit=limit)
        columns = self.get_roster_columns(roster_id)
        for member in members:
            values: Dict[str, str] = {}
            for col in columns:
                key = str(col.get("key", "")).strip()
                if not key:
                    continue
                values[key] = self._member_field_value(member, key)
            member["values"] = values
        return members

    def build_roster_column_list_objects(self, roster_id: str) -> List[Dict[str, Any]]:
        rid = str(roster_id or "").strip()
        if not rid:
            return []
        columns = self.get_roster_columns(rid)
        members = self.list_roster_members(rid, limit=100000)
        out: List[Dict[str, Any]] = []
        used_names: set[str] = set()
        for col in columns:
            key = str(col.get("key", "")).strip()
            label = str(col.get("label", "")).strip() or key
            if not key:
                continue
            seen: set[str] = set()
            items: List[Dict[str, str]] = []
            for member in members:
                value = str(member.get("values", {}).get(key, "")).strip()
                if not value or value in seen:
                    continue
                seen.add(value)
                items.append({"key": value, "label": value})
            list_type = "text"
            if items:
                all_numeric = True
                for item in items:
                    iv = str(item.get("key", "")).strip()
                    if not iv:
                        continue
                    try:
                        float(iv)
                    except Exception:
                        all_numeric = False
                        break
                if all_numeric:
                    list_type = "number"
            name = f"__名单列_{label}"
            if name in used_names:
                name = f"__名单列_{label}_{key}"
            used_names.add(name)
            out.append(
                {
                    "name": name,
                    "type": list_type,
                    "source": f"roster_auto:{rid}:{key}",
                    "readonly": True,
                    "roster_column_key": key,
                    "items": items,
                }
            )
        return out

    def _schema_uses_roster_repeat(self, schema: Dict[str, Any]) -> bool:
        questions = schema.get("questions", []) if isinstance(schema, dict) else []
        for question in questions:
            if str(question.get("repeat_from", "")).strip() == ROSTER_REPEAT_TOKEN:
                return True
        return False

    def _build_roster_repeat_items(self, roster_id: str, current_member_key: str = "") -> List[Dict[str, Any]]:
        rid = roster_id.strip()
        if not rid:
            return []
        members = self.db.list_roster_members(rid, limit=100000)
        columns = self.get_roster_columns(rid)
        items: List[Dict[str, Any]] = []
        seen: set[str] = set()
        current_key = current_member_key.strip()
        label_candidates = [str(c.get("key", "")).strip() for c in columns if str(c.get("key", "")).strip()]
        for member in members:
            key = str(member.get("member_key", "")).strip() or str(member.get("member_code", "")).strip()
            if not key or key in seen:
                continue
            seen.add(key)
            code = str(member.get("member_code", "")).strip()
            name = str(member.get("member_name", "")).strip()
            value_map: Dict[str, str] = {}
            extra = member.get("extra", {})
            if isinstance(extra, dict):
                for k, v in extra.items():
                    value_map[str(k)] = str(v)
            value_map["member_key"] = key
            if name:
                value_map["member_name"] = name
            if code:
                value_map["member_code"] = code
            if code and name:
                label = f"{code} - {name}"
            else:
                label = ""
                for lk in label_candidates:
                    vv = str(value_map.get(lk, "")).strip()
                    if vv:
                        label = vv
                        break
                if not label:
                    label = name or code or key
            items.append(
                {
                    "key": key,
                    "label": label,
                    "member_code": code,
                    "member_name": name,
                    "values": value_map,
                    "is_self": bool(current_key and key == current_key),
                }
            )
        return items

    def get_roster_repeat_items(self, questionnaire_id: str, current_member_key: str = "") -> List[Dict[str, Any]]:
        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            return []
        if not self._schema_uses_roster_repeat(questionnaire.get("schema", {})):
            return []
        roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
        return self._build_roster_repeat_items(roster_id, current_member_key=current_member_key)

    def add_roster_member(
        self,
        roster_id: str,
        member_name: str,
        member_code: str,
        member_key: str = "",
        tags: str = "",
        member_values: Optional[Dict[str, Any]] = None,
    ) -> None:
        if not roster_id.strip():
            raise ServiceError("名单不存在。")
        values = member_values if isinstance(member_values, dict) else {}
        columns = self.get_roster_columns(roster_id)
        normalized_values: Dict[str, str] = {}
        for idx, col in enumerate(columns, start=1):
            c_key = str(col.get("key", "")).strip() or f"field_{idx}"
            normalized_values[c_key] = str(values.get(c_key, "")).strip()
        if member_name.strip():
            normalized_values["member_name"] = member_name.strip()
        if member_code.strip():
            normalized_values["member_code"] = member_code.strip()
        key_field = next((str(c.get("key", "")).strip() for c in columns if bool(c.get("is_key", False))), "")
        key_from_values = normalized_values.get(key_field, "") if key_field else ""
        name = normalized_values.get("member_name", "")
        code = normalized_values.get("member_code", "")
        key = member_key.strip() or key_from_values or code or f"K{uuid.uuid4().hex[:10]}"
        extra: Dict[str, Any] = {}
        for k, v in normalized_values.items():
            if k in {"member_name", "member_code", "member_key"}:
                continue
            if v:
                extra[k] = v
        self.db.add_roster_member(roster_id, key, name, code, tags=tags.strip(), extra=extra)
        self.db.append_audit_log("roster_member_added", {"roster_id": roster_id, "member_key": key})

    def remove_roster_member(self, member_id: int) -> None:
        self.db.remove_roster_member(member_id)
        self.db.append_audit_log("roster_member_removed", {"member_id": member_id})

    def _infer_column(self, headers: List[str], candidates: List[str]) -> str:
        normalized_map = {h.strip().lower(): h for h in headers}
        for c in candidates:
            if c.strip().lower() in normalized_map:
                return normalized_map[c.strip().lower()]
        for h in headers:
            hl = h.strip().lower()
            if any(c in hl for c in candidates):
                return h
        return ""

    def _read_rows_from_file(self, file_path: Path) -> List[Dict[str, Any]]:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            with file_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                return [dict(row) for row in reader]
        if suffix in {".xlsx", ".xlsm"}:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                sheet = wb.active
                rows_iter = sheet.iter_rows(values_only=True)
                headers_row = next(rows_iter, None)
                if not headers_row:
                    return []
                headers = [str(h).strip() if h is not None else "" for h in headers_row]
                rows: List[Dict[str, Any]] = []
                for row in rows_iter:
                    item: Dict[str, Any] = {}
                    for idx, cell in enumerate(row):
                        if idx >= len(headers):
                            continue
                        key = headers[idx]
                        if not key:
                            continue
                        item[key] = "" if cell is None else str(cell)
                    rows.append(item)
                return rows
            finally:
                wb.close()
        raise ServiceError("仅支持 CSV 或 Excel(xlsx/xlsm) 导入。")

    def import_roster_file(self, roster_id: str, file_path: Path, replace_all: bool = False) -> Dict[str, int]:
        if not file_path.exists():
            raise ServiceError("导入文件不存在。")
        rows = self._read_rows_from_file(file_path)
        if not rows:
            raise ServiceError("导入文件为空或无有效数据。")
        roster_columns = self.get_roster_columns(roster_id)
        headers = list(rows[0].keys())
        column_map: Dict[str, str] = {}
        for idx, col in enumerate(roster_columns, start=1):
            c_key = str(col.get("key", "")).strip() or f"field_{idx}"
            c_label = str(col.get("label", "")).strip()
            found = self._infer_column(
                headers,
                [c_label, c_key, c_key.replace("_", ""), c_key.lower()],
            )
            if found:
                column_map[c_key] = found
        key_col = self._infer_column(headers, ["唯一标识", "member_key", "key", "账号"])
        if not key_col:
            key_field = next((str(c.get("key", "")).strip() for c in roster_columns if bool(c.get("is_key", False))), "")
            key_col = column_map.get(key_field, "")
        members: List[Dict[str, Any]] = []
        for row in rows:
            values: Dict[str, str] = {}
            for col in roster_columns:
                c_key = str(col.get("key", "")).strip()
                if not c_key:
                    continue
                src = column_map.get(c_key, "")
                values[c_key] = str(row.get(src, "")).strip() if src else ""

            name = str(values.get("member_name", "")).strip()
            code = str(values.get("member_code", "")).strip()
            tags = str(row.get("tags", "")).strip()
            key = str(row.get(key_col, "")).strip() if key_col else ""
            if not key:
                key_field = next((str(c.get("key", "")).strip() for c in roster_columns if bool(c.get("is_key", False))), "")
                if key_field:
                    key = str(values.get(key_field, "")).strip()
            if not key:
                key = code or f"K{uuid.uuid4().hex[:10]}"

            extra: Dict[str, Any] = {}
            for c_key, c_val in values.items():
                if c_key in {"member_name", "member_code", "member_key"}:
                    continue
                if c_val != "":
                    extra[c_key] = c_val
            used_headers = {h for h in column_map.values() if h}
            if key_col:
                used_headers.add(key_col)
            used_headers.add("tags")
            for hk, hv in row.items():
                hks = str(hk).strip()
                if not hks or hks in used_headers:
                    continue
                text = str(hv).strip()
                if text:
                    extra[hks] = text
            if not key and not code and not name and not any(str(v).strip() for v in extra.values()):
                continue
            members.append(
                {
                    "member_key": key,
                    "member_code": code,
                    "member_name": name,
                    "tags": tags,
                    "extra": extra,
                }
            )
        if not members:
            raise ServiceError("未识别到有效成员数据。")
        result = self.db.upsert_roster_members(roster_id=roster_id, members=members, replace_all=replace_all)
        self.db.append_audit_log(
            "roster_imported",
            {
                "roster_id": roster_id,
                "file": str(file_path),
                "replace_all": bool(replace_all),
                "inserted": result["inserted"],
                "updated": result["updated"],
            },
        )
        return result

    def _questionnaire_collect_fields(self, questionnaire: Dict[str, Any]) -> List[Dict[str, str]]:
        fields = questionnaire.get("identity_fields", {})
        if not isinstance(fields, dict):
            fields = {}
        collect_fields = self._normalize_collect_fields(fields.get("collect_fields"))
        if collect_fields:
            return collect_fields
        legacy = self._legacy_collect_fields(fields)
        if legacy:
            return legacy
        auth_mode = str(questionnaire.get("auth_mode", "open")).strip().lower()
        if auth_mode == "roster_name_code":
            return [{"key": "member_code", "label": "编号"}, {"key": "member_name", "label": "姓名"}]
        if auth_mode == "roster_code":
            return [{"key": "member_code", "label": "编号"}]
        if auth_mode == "roster_fields":
            roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
            cols = self.get_roster_columns(roster_id)
            return [{"key": c["key"], "label": c["label"]} for c in cols if str(c.get("key", "")).strip()]
        return []

    def _questionnaire_allow_same_device_repeat(self, questionnaire: Dict[str, Any]) -> bool:
        fields = questionnaire.get("identity_fields", {})
        if not isinstance(fields, dict):
            return False
        return bool(fields.get("allow_same_device_repeat", False))

    def _member_identity_value(self, member: Dict[str, Any], key: str) -> str:
        value = self._member_field_value(member, key)
        return str(value or "").strip()

    def _normalize_identity_data_input(
        self,
        collect_fields: List[Dict[str, str]],
        identity_data: Optional[Dict[str, Any]],
        member_code: str,
        member_name: str,
    ) -> Dict[str, str]:
        source = identity_data if isinstance(identity_data, dict) else {}
        result: Dict[str, str] = {}
        for field in collect_fields:
            key = str(field.get("key", "")).strip()
            if not key:
                continue
            result[key] = str(source.get(key, "")).strip()
        if member_code.strip():
            result["member_code"] = member_code.strip()
        if member_name.strip():
            result["member_name"] = member_name.strip()
        return result

    def verify_submission_identity(
        self,
        questionnaire_id: str,
        member_code: str = "",
        member_name: str = "",
        identity_data: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            raise ServiceError("问卷不存在。")
        auth_mode = str(questionnaire.get("auth_mode", "open")).strip().lower() or "open"
        roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
        collect_fields = self._questionnaire_collect_fields(questionnaire)
        provided_identity = self._normalize_identity_data_input(
            collect_fields=collect_fields,
            identity_data=identity_data,
            member_code=member_code,
            member_name=member_name,
        )
        for field in collect_fields:
            k = str(field.get("key", "")).strip()
            label = str(field.get("label", "")).strip() or k
            if not k:
                continue
            if not str(provided_identity.get(k, "")).strip():
                raise ServiceError(f"请先填写“{label}”。")

        if auth_mode == "open":
            return {
                "auth_required": False,
                "auth_token": "",
                "member": {
                    "member_key": "",
                    "name": str(provided_identity.get("member_name", "")).strip(),
                    "code": str(provided_identity.get("member_code", "")).strip(),
                    "identity_data": provided_identity,
                },
            }
        if not roster_id:
            raise ServiceError("问卷未绑定名单，无法校验身份。")

        if auth_mode == "roster_name_code":
            lookup_fields = {
                "member_code": str(provided_identity.get("member_code", "")).strip(),
                "member_name": str(provided_identity.get("member_name", "")).strip(),
            }
        elif auth_mode == "roster_code":
            lookup_fields = {"member_code": str(provided_identity.get("member_code", "")).strip()}
        else:
            lookup_fields = {}
            if collect_fields:
                for field in collect_fields:
                    key = str(field.get("key", "")).strip()
                    if not key:
                        continue
                    lookup_fields[key] = str(provided_identity.get(key, "")).strip()
            else:
                lookup_fields = provided_identity

        member = self.db.find_roster_member_by_fields(
            roster_id=roster_id,
            fields=lookup_fields,
        )
        if not member:
            raise ServiceError("身份验证失败：不在允许名单中。")

        identity_out = dict(provided_identity)
        for field in collect_fields:
            k = str(field.get("key", "")).strip()
            if not k:
                continue
            if not str(identity_out.get(k, "")).strip():
                identity_out[k] = self._member_identity_value(member, k)
        for col in self.get_roster_columns(roster_id):
            k = str(col.get("key", "")).strip()
            if not k:
                continue
            if not str(identity_out.get(k, "")).strip():
                value = self._member_identity_value(member, k)
                if value:
                    identity_out[k] = value

        token = secrets.token_urlsafe(28)
        expires_at = (datetime.now(timezone.utc) + timedelta(hours=2)).isoformat()
        self.db.create_auth_session(
            token=token,
            questionnaire_id=questionnaire_id,
            roster_id=roster_id,
            member_key=member["member_key"],
            expires_at=expires_at,
        )
        self.db.append_audit_log(
            "identity_verified",
            {
                "questionnaire_id": questionnaire_id,
                "roster_id": roster_id,
                "member_key": member["member_key"],
            },
        )
        return {
            "auth_required": True,
            "auth_token": token,
            "member": {
                "member_key": member.get("member_key", ""),
                "name": member.get("member_name", ""),
                "code": member.get("member_code", ""),
                "identity_data": identity_out,
            },
        }

    def _consume_verified_token(self, questionnaire: Dict[str, Any], auth_token: str) -> Dict[str, Any]:
        auth_mode = questionnaire.get("auth_mode", "open")
        if auth_mode == "open":
            return {"roster_id": "", "member_key": "", "member_code": "", "member_name": "", "identity_data": {}}
        token = auth_token.strip()
        if not token:
            raise ServiceError("需要先通过身份验证。")
        self.db.purge_expired_auth_sessions(utc_now())
        session = self.db.get_auth_session(token)
        if not session:
            raise ServiceError("身份会话无效，请重新验证。")
        if session.get("questionnaire_id") != questionnaire["id"]:
            raise ServiceError("身份会话与问卷不匹配。")
        if int(session.get("used", 0)) == 1:
            raise ServiceError("身份会话已使用，请重新验证。")
        if parse_iso(session["expires_at"]) < datetime.now(timezone.utc):
            raise ServiceError("身份会话已过期，请重新验证。")

        roster_id = str(session.get("roster_id", "")).strip()
        member_key = str(session.get("member_key", "")).strip()
        member_name = ""
        member_code = ""
        identity_data: Dict[str, str] = {}
        if roster_id and member_key:
            members = self.db.list_roster_members(roster_id, limit=50000)
            found = next((m for m in members if m.get("member_key") == member_key), None)
            if found:
                member_name = str(found.get("member_name", ""))
                member_code = str(found.get("member_code", ""))
                for col in self.get_roster_columns(roster_id):
                    fkey = str(col.get("key", "")).strip()
                    if not fkey:
                        continue
                    value = self._member_field_value(found, fkey)
                    if value:
                        identity_data[fkey] = value
                for field in self._questionnaire_collect_fields(questionnaire):
                    fkey = str(field.get("key", "")).strip()
                    if not fkey:
                        continue
                    if fkey not in identity_data:
                        value = self._member_field_value(found, fkey)
                        if value:
                            identity_data[fkey] = value
        return {
            "token": token,
            "roster_id": roster_id,
            "member_key": member_key,
            "member_name": member_name,
            "member_code": member_code,
            "identity_data": identity_data,
        }

    def _resolve_current_member_key(
        self,
        questionnaire: Dict[str, Any],
        verified_member_key: str,
        respondent_code: str,
        respondent_name: str,
        respondent_identity: Optional[Dict[str, Any]] = None,
    ) -> str:
        key = verified_member_key.strip()
        if key:
            return key
        roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
        if not roster_id:
            return ""
        code = respondent_code.strip()
        name = respondent_name.strip()
        auth_mode = str(questionnaire.get("auth_mode", "open")).strip().lower()
        identity = respondent_identity if isinstance(respondent_identity, dict) else {}
        if auth_mode == "roster_fields":
            lookup: Dict[str, str] = {}
            for field in self._questionnaire_collect_fields(questionnaire):
                fkey = str(field.get("key", "")).strip()
                if not fkey:
                    continue
                value = str(identity.get(fkey, "")).strip()
                if not value and fkey == "member_code":
                    value = code
                if not value and fkey == "member_name":
                    value = name
                if value:
                    lookup[fkey] = value
            member = self.db.find_roster_member_by_fields(roster_id=roster_id, fields=lookup)
            if not member:
                return ""
            return str(member.get("member_key", "")).strip()

        if not code and not name:
            return ""
        mode = "roster_name_code" if (code and name) else "roster_code"
        member = self.db.find_roster_member(
            roster_id=roster_id,
            mode=mode,
            member_code=code,
            member_name=name,
        )
        if not member:
            return ""
        return str(member.get("member_key", "")).strip()

    def _inject_schema_list_sources(
        self,
        schema: Dict[str, Any],
        answers_for_validate: Dict[str, Any],
        current_member_key: str = "",
    ) -> None:
        meta = schema.get("meta", {}) if isinstance(schema, dict) else {}
        if not isinstance(meta, dict):
            return
        list_objects = meta.get("list_objects", [])
        if not isinstance(list_objects, list):
            return
        current_key = current_member_key.strip()
        for obj in list_objects:
            if not isinstance(obj, dict):
                continue
            name = str(obj.get("name", "")).strip()
            if not name:
                continue
            token = f"__list__:{name}"
            items = obj.get("items", [])
            if not isinstance(items, list):
                continue
            normalized_items: List[Dict[str, Any]] = []
            seen: set[str] = set()
            for raw in items:
                if isinstance(raw, dict):
                    key = str(raw.get("key", "")).strip() or str(raw.get("value", "")).strip()
                    label = str(raw.get("label", "")).strip() or key
                else:
                    key = str(raw).strip()
                    label = key
                if not key or key in seen:
                    continue
                seen.add(key)
                normalized_items.append(
                    {
                        "key": key,
                        "label": label,
                        "is_self": bool(current_key and key == current_key),
                    }
                )
            if normalized_items:
                answers_for_validate[token] = normalized_items

    def _inject_legacy_repeat_source(
        self,
        schema: Dict[str, Any],
        answers_for_validate: Dict[str, Any],
    ) -> None:
        # 兼容历史问卷：旧版使用 __roster_members__ 作为循环来源。
        if ROSTER_REPEAT_TOKEN in answers_for_validate:
            return
        questions = schema.get("questions", []) if isinstance(schema, dict) else []
        if not isinstance(questions, list):
            return
        items: List[Dict[str, Any]] = []
        seen: set[str] = set()
        for q in questions:
            if not isinstance(q, dict):
                continue
            if str(q.get("repeat_from", "")).strip() != ROSTER_REPEAT_TOKEN:
                continue
            qid = str(q.get("id", "")).strip()
            if not qid:
                continue
            raw = answers_for_validate.get(qid)
            if not isinstance(raw, dict):
                continue
            for k in raw.keys():
                key = str(k).strip()
                if not key or key in seen:
                    continue
                seen.add(key)
                items.append({"key": key, "label": key, "is_self": False})
        if items:
            answers_for_validate[ROSTER_REPEAT_TOKEN] = items

    def _vote_file_path(self, questionnaire_id: str, submission_id: str) -> Path:
        return self.paths.votes_dir / questionnaire_id / f"{submission_id}.vote"

    def submit_response(
        self,
        questionnaire_id: str,
        answers: Dict[str, Any],
        respondent_name: str = "",
        respondent_code: str = "",
        respondent_identity: Optional[Dict[str, Any]] = None,
        client_token: str = "",
        source: str = "lan",
        relation_type: str = "",
        target_label: str = "",
        auth_token: str = "",
    ) -> SubmissionResult:
        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            raise ServiceError("问卷不存在。")
        if questionnaire.get("status") != "active":
            raise ServiceError("问卷未开放。")

        verified = self._consume_verified_token(questionnaire, auth_token)
        identity_mode = self._normalize_identity_mode(questionnaire["identity_mode"])
        anonymous = False
        collect_fields = self._questionnaire_collect_fields(questionnaire)
        raw_identity = respondent_identity if isinstance(respondent_identity, dict) else {}
        if respondent_name.strip():
            raw_identity["member_name"] = respondent_name.strip()
        if respondent_code.strip():
            raw_identity["member_code"] = respondent_code.strip()
        identity_data: Dict[str, str] = {}
        verified_identity = verified.get("identity_data", {}) if isinstance(verified.get("identity_data", {}), dict) else {}
        for field in collect_fields:
            key = str(field.get("key", "")).strip()
            label = str(field.get("label", "")).strip() or key
            if not key:
                continue
            value = str(raw_identity.get(key, "")).strip()
            if not value:
                value = str(verified_identity.get(key, "")).strip()
            if not value and key == "member_name":
                value = str(verified.get("member_name", "")).strip()
            if not value and key == "member_code":
                value = str(verified.get("member_code", "")).strip()
            if not value:
                raise ServiceError(f"请先填写“{label}”。")
            identity_data[key] = value
        for key, raw_value in verified_identity.items():
            k = str(key).strip()
            if not k or k in identity_data:
                continue
            value = str(raw_value).strip()
            if value:
                identity_data[k] = value

        name = str(identity_data.get("member_name", "")).strip() or respondent_name.strip()
        code = str(identity_data.get("member_code", "")).strip() or respondent_code.strip()

        # 名单验证通过时，若页面未采集姓名/编号，用名单中的字段补充去重信息。
        verified_member_key = str(verified.get("member_key", "")).strip()
        verified_code = str(verified.get("member_code", "")).strip()
        verified_name = str(verified.get("member_name", "")).strip()
        identity_dedupe_key = self._identity_dedupe_key(identity_data)
        dedupe_code = code or verified_code or identity_dedupe_key
        if not name and verified_name:
            name = verified_name
        if not code and verified_code:
            code = verified_code

        schema = questionnaire["schema"]
        current_member_key = self._resolve_current_member_key(
            questionnaire=questionnaire,
            verified_member_key=verified_member_key,
            respondent_code=code,
            respondent_name=name,
            respondent_identity=identity_data,
        )
        answers_for_validate = dict(answers)
        self._inject_schema_list_sources(
            schema=schema,
            answers_for_validate=answers_for_validate,
            current_member_key=current_member_key,
        )
        self._inject_legacy_repeat_source(schema=schema, answers_for_validate=answers_for_validate)

        ok, errors, cleaned_answers = validate_answers(schema, answers_for_validate)
        if not ok:
            raise ServiceError("；".join(errors))
        sql_rule_errors = self._evaluate_sql_validation_rules(
            questionnaire=questionnaire,
            answers=cleaned_answers,
            respondent_name=name,
            respondent_code=code,
            respondent_identity=identity_data,
            current_member_key=current_member_key,
        )
        if sql_rule_errors:
            raise ServiceError("；".join(sql_rule_errors))

        allow_repeat = questionnaire["allow_repeat"]
        allow_same_device_repeat = self._questionnaire_allow_same_device_repeat(questionnaire)
        dedupe_token = client_token.strip()
        dedupe_member_key = verified_member_key or current_member_key or identity_dedupe_key
        if not allow_repeat and self.db.detect_duplicate(
            questionnaire_id=questionnaire_id,
            client_token="" if allow_same_device_repeat else dedupe_token,
            respondent_code=dedupe_code,
            verified_member_key=dedupe_member_key,
        ):
            if allow_same_device_repeat:
                raise ServiceError("当前身份已提交，不能重复作答。")
            raise ServiceError("当前身份或设备已提交，不能重复作答。")

        submission_id = f"S{uuid.uuid4().hex[:14]}"
        payload = {
            "submission_id": submission_id,
            "questionnaire_id": questionnaire_id,
            "questionnaire_title": questionnaire["title"],
            "questionnaire_version": int(questionnaire.get("current_version", 1)),
            "submitted_from": source,
            "identity_mode": identity_mode,
            "respondent": {
                "name": "" if anonymous else name,
                "code": "" if anonymous else code,
                "identity_data": identity_data,
                "anonymous": anonymous,
                "client_token": dedupe_token,
            },
            "verified": {
                "roster_id": str(verified.get("roster_id", "")),
                "member_key": verified_member_key,
            },
            "context": {
                "relation_type": relation_type,
                "target_label": target_label,
                "current_member_key": dedupe_member_key,
            },
            "answers": cleaned_answers,
        }

        envelope = self.crypto.encrypt_payload(payload, source=source)
        vote_path = self._vote_file_path(questionnaire_id, submission_id)
        self.crypto.save_vote_file(envelope, vote_path)

        self.db.save_submission_meta(
            submission_id=submission_id,
            questionnaire_id=questionnaire_id,
            questionnaire_version=int(questionnaire.get("current_version", 1)),
            respondent_name="" if anonymous else name,
            respondent_code="" if anonymous else code,
            anonymous=anonymous,
            source=source,
            vote_file=str(vote_path),
            client_token=dedupe_token,
            session_label=relation_type,
            target_label=target_label,
            roster_id=str(verified.get("roster_id", "")),
            verified_member_key=dedupe_member_key,
        )

        if verified.get("token"):
            self.db.consume_auth_session(str(verified["token"]))
        self.db.append_audit_log(
            "submission_saved",
            {"questionnaire_id": questionnaire_id, "submission_id": submission_id, "source": source},
        )
        return SubmissionResult(submission_id=submission_id, vote_path=vote_path)

    def list_submissions(self, questionnaire_id: Optional[str] = None) -> List[Dict[str, Any]]:
        return self.db.list_submissions(questionnaire_id=questionnaire_id)

    def reject_submission(self, submission_id: str) -> None:
        sid = str(submission_id or "").strip()
        if not sid:
            raise ServiceError("票据不存在。")
        row = self.db.get_submission(sid)
        if not row:
            raise ServiceError("票据不存在。")
        vote_file = Path(str(row.get("vote_file", "")).strip()) if str(row.get("vote_file", "")).strip() else None
        self.db.delete_submission(sid)
        if vote_file is not None:
            try:
                if vote_file.exists():
                    vote_file.unlink()
            except Exception:
                pass
        self.db.append_audit_log(
            "submission_rejected",
            {"submission_id": sid, "questionnaire_id": str(row.get("questionnaire_id", ""))},
        )

    def decrypt_submission_payloads(self, questionnaire_id: str) -> List[Dict[str, Any]]:
        if not self.crypto.unlocked:
            raise ServiceError("管理员密钥尚未解锁。")
        payloads: List[Dict[str, Any]] = []
        for item in self.list_submissions(questionnaire_id):
            vote_path = Path(item["vote_file"])
            if not vote_path.exists():
                continue
            try:
                payload = self.crypto.decrypt_vote_file(vote_path)
                payloads.append(payload)
            except VoteCryptoError:
                continue
        return payloads

    def list_sql_views(self, questionnaire_id: str) -> List[Dict[str, Any]]:
        qid = str(questionnaire_id or "").strip()
        if not qid:
            return []
        questionnaire = self.get_questionnaire(qid)
        if questionnaire:
            template_key = self._questionnaire_template_key(questionnaire)
            if template_key:
                self._sync_template_sql_views_to_questionnaire(qid, template_key, overwrite=False)
        return self.db.list_sql_views(qid)

    def save_sql_view(self, questionnaire_id: str, name: str, sql_text: str) -> int:
        qid = str(questionnaire_id or "").strip()
        view_name = str(name or "").strip()
        sql = self._normalize_query_script(sql_text)
        if not qid:
            raise ServiceError("问卷不存在。")
        if not self.get_questionnaire(qid):
            raise ServiceError("问卷不存在。")
        if not view_name:
            raise ServiceError("查询名称不能为空。")
        if len(view_name) > 64:
            raise ServiceError("查询名称过长（最多64个字符）。")
        view_id = self.db.save_sql_view(qid, view_name, sql)
        questionnaire = self.get_questionnaire(qid)
        template_key = self._questionnaire_template_key(questionnaire)
        if template_key:
            self.db.save_template_sql_view(template_key, view_name, sql)
        self.db.append_audit_log(
            "sql_view_saved",
            {"questionnaire_id": qid, "name": view_name, "template_key": template_key},
        )
        return view_id

    def remove_sql_view(self, view_id: int) -> None:
        if int(view_id) <= 0:
            raise ServiceError("查询模板不存在。")
        row = self.db.get_sql_view(int(view_id))
        if not row:
            raise ServiceError("查询模板不存在。")
        qid = str(row.get("questionnaire_id", "")).strip()
        name = str(row.get("name", "")).strip()
        self.db.remove_sql_view(int(view_id))
        template_key = ""
        questionnaire = self.get_questionnaire(qid) if qid else None
        if questionnaire:
            template_key = self._questionnaire_template_key(questionnaire)
            if template_key and name:
                self.db.remove_template_sql_view(template_key, name)
        self.db.append_audit_log(
            "sql_view_removed",
            {"view_id": int(view_id), "questionnaire_id": qid, "name": name, "template_key": template_key},
        )

    def validate_live_rule_sql(self, sql_text: str) -> str:
        return self._normalize_live_rule_sql(sql_text)

    def _split_sql_statements(self, sql_text: str) -> List[str]:
        text = str(sql_text or "").strip()
        if not text:
            return []
        stmts: List[str] = []
        buf: List[str] = []
        quote: Optional[str] = None
        line_comment = False
        block_comment = False
        i = 0
        while i < len(text):
            ch = text[i]
            nxt = text[i + 1] if i + 1 < len(text) else ""
            if quote:
                buf.append(ch)
                if ch == quote:
                    # SQL 字符串转义：'' 或 ""
                    if i + 1 < len(text) and text[i + 1] == quote:
                        buf.append(text[i + 1])
                        i += 1
                    else:
                        quote = None
            elif line_comment:
                buf.append(ch)
                if ch in {"\n", "\r"}:
                    line_comment = False
            elif block_comment:
                buf.append(ch)
                if ch == "*" and nxt == "/":
                    buf.append(nxt)
                    i += 1
                    block_comment = False
            else:
                if ch in {"'", '"'}:
                    quote = ch
                    buf.append(ch)
                elif ch == "-" and nxt == "-":
                    buf.append(ch)
                    buf.append(nxt)
                    i += 1
                    line_comment = True
                elif ch == "/" and nxt == "*":
                    buf.append(ch)
                    buf.append(nxt)
                    i += 1
                    block_comment = True
                elif ch == ";":
                    part = "".join(buf).strip()
                    if part:
                        stmts.append(part)
                    buf = []
                else:
                    buf.append(ch)
            i += 1
        tail = "".join(buf).strip()
        if tail:
            stmts.append(tail)
        return stmts

    def _strip_leading_sql_comments(self, text: str) -> str:
        s = str(text or "")
        idx = 0
        n = len(s)
        while idx < n:
            while idx < n and s[idx].isspace():
                idx += 1
            if idx + 1 < n and s[idx : idx + 2] == "--":
                nl = s.find("\n", idx + 2)
                if nl < 0:
                    return ""
                idx = nl + 1
                continue
            if idx + 1 < n and s[idx : idx + 2] == "/*":
                end = s.find("*/", idx + 2)
                if end < 0:
                    return ""
                idx = end + 2
                continue
            break
        return s[idx:].strip()

    def _mask_sql_literals_and_comments(self, text: str) -> str:
        s = str(text or "")
        out: List[str] = []
        i = 0
        n = len(s)
        quote: Optional[str] = None
        while i < n:
            ch = s[i]
            nxt = s[i + 1] if i + 1 < n else ""
            if quote:
                if ch == quote:
                    if i + 1 < n and s[i + 1] == quote:
                        out.append("  ")
                        i += 2
                        continue
                    quote = None
                out.append(" ")
                i += 1
                continue
            if ch in {"'", '"'}:
                quote = ch
                out.append(" ")
                i += 1
                continue
            if ch == "-" and nxt == "-":
                out.append("  ")
                i += 2
                while i < n and s[i] not in {"\n", "\r"}:
                    out.append(" ")
                    i += 1
                continue
            if ch == "/" and nxt == "*":
                out.append("  ")
                i += 2
                while i + 1 < n and not (s[i] == "*" and s[i + 1] == "/"):
                    out.append(" ")
                    i += 1
                if i + 1 < n:
                    out.append("  ")
                    i += 2
                continue
            out.append(ch)
            i += 1
        return "".join(out)

    def _remove_sql_comments(self, text: str) -> str:
        s = str(text or "")
        out: List[str] = []
        i = 0
        n = len(s)
        quote: Optional[str] = None
        while i < n:
            ch = s[i]
            nxt = s[i + 1] if i + 1 < n else ""
            if quote:
                out.append(ch)
                if ch == quote:
                    if i + 1 < n and s[i + 1] == quote:
                        out.append(s[i + 1])
                        i += 2
                        continue
                    quote = None
                i += 1
                continue
            if ch in {"'", '"'}:
                quote = ch
                out.append(ch)
                i += 1
                continue
            if ch == "-" and nxt == "-":
                i += 2
                while i < n and s[i] not in {"\n", "\r"}:
                    i += 1
                continue
            if ch == "/" and nxt == "*":
                i += 2
                while i + 1 < n and not (s[i] == "*" and s[i + 1] == "/"):
                    i += 1
                if i + 1 < n:
                    i += 2
                continue
            out.append(ch)
            i += 1
        return "".join(out)

    def _normalize_select_statement(self, sql_text: str) -> str:
        text = str(sql_text or "").strip()
        if not text:
            raise ServiceError("SQL 不能为空。")
        normalized_head = self._strip_leading_sql_comments(text)
        lowered = normalized_head.lower()
        if not (lowered.startswith("select") or lowered.startswith("with")):
            raise ServiceError("仅允许 SELECT 查询语句。")
        sanitized = self._mask_sql_literals_and_comments(text)
        blocked = re.compile(
            r"\b(insert|update|delete|drop|alter|create|truncate|attach|detach|pragma|vacuum)\b",
            re.IGNORECASE,
        )
        if blocked.search(sanitized):
            raise ServiceError("SQL 包含禁止关键字，仅支持只读查询。")
        return text

    def _normalize_query_script(self, sql_text: str) -> str:
        statements = self._split_sql_statements(sql_text)
        if not statements:
            raise ServiceError("SQL 不能为空。")
        normalized = [self._normalize_select_statement(stmt) for stmt in statements]
        return ";\n".join(normalized)

    def _normalize_live_rule_op(self, raw_op: str) -> str:
        op = str(raw_op or "").strip().lower()
        mapping = {
            "=": "equals",
            "==": "equals",
            "eq": "equals",
            "equals": "equals",
            "!=": "not_equals",
            "<>": "not_equals",
            "neq": "not_equals",
            "not_equals": "not_equals",
            ">": "gt",
            "gt": "gt",
            ">=": "gte",
            "gte": "gte",
            "<": "lt",
            "lt": "lt",
            "<=": "lte",
            "lte": "lte",
            "between": "between",
            "not_between": "not_between",
        }
        return mapping.get(op, op)

    def _parse_rule_numeric(self, raw_value: Any, label: str) -> float:
        if raw_value is None:
            raise ServiceError(f"{label}不能为空。")
        text = str(raw_value).strip()
        if not text:
            raise ServiceError(f"{label}不能为空。")
        try:
            return float(text)
        except Exception as exc:
            raise ServiceError(f"{label}必须是数字。") from exc

    def _live_rule_compare(self, actual: float, op: str, value: float, value2: Optional[float] = None) -> bool:
        if op == "equals":
            return actual == value
        if op == "not_equals":
            return actual != value
        if op == "gt":
            return actual > value
        if op == "gte":
            return actual >= value
        if op == "lt":
            return actual < value
        if op == "lte":
            return actual <= value
        if op == "between":
            if value2 is None:
                return False
            low = min(value, value2)
            high = max(value, value2)
            return low <= actual <= high
        if op == "not_between":
            if value2 is None:
                return False
            low = min(value, value2)
            high = max(value, value2)
            return not (low <= actual <= high)
        return False

    def _live_rule_allowed_tables(self) -> List[str]:
        return [
            "submissions",
            "identity_kv",
            "question_defs",
            "question_options",
            "answers",
            "answer_options",
            "v_scores",
            "v_text_answers",
            "v_answers_enriched",
            "v_answer_options_enriched",
            "v_identity_enriched",
            "v_submissions_identity",
            "v_scores_enriched",
            "v_scores_identity",
        ]

    def _normalize_live_rule_sql(self, sql_text: str) -> str:
        statements = self._split_sql_statements(sql_text)
        if len(statements) != 1:
            raise ServiceError("联合规则 SQL 只能填写 1 条 SELECT 语句。")
        sql_raw = self._normalize_select_statement(statements[0]).strip().rstrip(";")
        sql = self._remove_sql_comments(sql_raw).strip().rstrip(";")
        if not sql:
            raise ServiceError("联合规则 SQL 不能为空。")
        return sql

    def live_rule_auto_filter_suffix(self) -> str:
        return "当前联合规则模型仅包含“当前答卷”，无需额外追加 submission_id 条件。"

    def _compose_live_rule_sql(self, normalized_sql: str, submission_id: str = "__current__") -> str:
        _ = submission_id
        return str(normalized_sql or "").strip().rstrip(";")

    def _sql_quote_literal(self, value: str) -> str:
        return "'" + str(value).replace("'", "''") + "'"

    def _sql_quote_ident(self, value: str) -> str:
        return '"' + str(value).replace('"', '""') + '"'

    def _identity_alias_seed(self, field_key: str, field_label: str = "") -> str:
        key = str(field_key or "").strip()
        label = str(field_label or "").strip()
        text = label or key

        phrase_map = {
            "member_name": "xing_ming",
            "member_code": "bian_hao",
            "member_key": "wei_yi_biao_shi",
            "姓名": "xing_ming",
            "学号": "xue_hao",
            "工号": "gong_hao",
            "编号": "bian_hao",
            "昵称": "ni_cheng",
            "班级": "ban_ji",
            "部门": "bu_men",
            "组别": "zu_bie",
            "角色": "jue_se",
            "学院": "xue_yuan",
            "专业": "zhuan_ye",
            "手机号": "shou_ji_hao",
            "电话": "dian_hua",
            "单位": "dan_wei",
            "岗位": "gang_wei",
        }
        if label in phrase_map:
            return phrase_map[label]
        if key in phrase_map:
            return phrase_map[key]

        text_ascii = re.sub(r"[^a-zA-Z0-9]+", "_", text.lower()).strip("_")
        if text_ascii:
            if text_ascii[0].isdigit():
                return f"f_{text_ascii}"
            return text_ascii

        char_map = {
            "姓": "xing",
            "名": "ming",
            "学": "xue",
            "号": "hao",
            "工": "gong",
            "编": "bian",
            "昵": "ni",
            "称": "cheng",
            "班": "ban",
            "级": "ji",
            "部": "bu",
            "门": "men",
            "组": "zu",
            "别": "bie",
            "角": "jue",
            "色": "se",
            "院": "yuan",
            "专": "zhuan",
            "业": "ye",
            "手": "shou",
            "机": "ji",
            "电": "dian",
            "话": "hua",
            "单": "dan",
            "位": "wei",
            "岗": "gang",
            "证": "zheng",
            "身": "shen",
            "份": "fen",
            "唯": "wei",
            "一": "yi",
            "标": "biao",
            "识": "shi",
        }
        parts: List[str] = []
        for ch in text:
            if ch in char_map:
                parts.append(char_map[ch])
        if parts:
            seed = "_".join(parts)
            seed = re.sub(r"_+", "_", seed).strip("_")
            if seed:
                return seed
        return "field"

    def _identity_alias_map(
        self,
        field_defs: List[Dict[str, str]],
        extra_keys: Optional[List[str]] = None,
    ) -> Dict[str, str]:
        label_map: Dict[str, str] = {}
        ordered_keys: List[str] = []
        seen_keys: set[str] = set()
        for item in field_defs:
            key = str(item.get("key", "")).strip()
            if not key:
                continue
            label_map[key] = str(item.get("label", "")).strip() or key
            if key not in seen_keys:
                seen_keys.add(key)
                ordered_keys.append(key)
        for key in extra_keys or []:
            k = str(key or "").strip()
            if not k or k in seen_keys:
                continue
            seen_keys.add(k)
            ordered_keys.append(k)

        result: Dict[str, str] = {}
        used_aliases: set[str] = set()
        for key in ordered_keys:
            seed = self._identity_alias_seed(key, label_map.get(key, key))
            alias = seed
            suffix = 2
            while alias in used_aliases:
                alias = f"{seed}_{suffix}"
                suffix += 1
            used_aliases.add(alias)
            result[key] = alias
        return result

    def _list_object_value_maps(self, schema_meta: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
        meta = schema_meta if isinstance(schema_meta, dict) else {}
        list_objects = meta.get("list_objects", [])
        out: Dict[str, Dict[str, str]] = {}
        if not isinstance(list_objects, list):
            return out
        for obj in list_objects:
            if not isinstance(obj, dict):
                continue
            name = str(obj.get("name", "")).strip()
            if not name:
                continue
            items = obj.get("items", [])
            if not isinstance(items, list):
                continue
            key_map: Dict[str, str] = {}
            for raw in items:
                if isinstance(raw, dict):
                    key = str(raw.get("key", "")).strip() or str(raw.get("value", "")).strip()
                    label = str(raw.get("label", "")).strip() or key
                else:
                    key = str(raw).strip()
                    label = key
                if not key:
                    continue
                key_map[key] = label
            if key_map:
                out[name] = key_map
        return out

    def _resolve_repeat_item_value(
        self,
        question: Dict[str, Any],
        repeat_key: str,
        schema_meta: Optional[Dict[str, Any]] = None,
    ) -> str:
        key = str(repeat_key or "").strip()
        if not key:
            return ""
        repeat_from = str(question.get("repeat_from", "")).strip()
        if not repeat_from:
            return key
        if repeat_from.startswith("__list__:"):
            list_name = repeat_from.split(":", 1)[1].strip()
            maps = self._list_object_value_maps(schema_meta)
            label = str(maps.get(list_name, {}).get(key, "")).strip()
            return label or key
        return key

    def _query_model_tables(self) -> List[Dict[str, Any]]:
        return [
            {
                "name": "submissions",
                "desc": "每份票据一行（归票主表）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("questionnaire_id", "TEXT", "问卷ID"),
                    ("questionnaire_version", "INTEGER", "问卷版本"),
                    ("submitted_at", "TEXT", "提交时间"),
                    ("source", "TEXT", "来源: lan/offline"),
                    ("anonymous", "INTEGER", "历史兼容字段：当前系统固定为0"),
                    ("respondent_name", "TEXT", "填写者姓名"),
                    ("respondent_code", "TEXT", "填写者编号"),
                    ("verified_member_key", "TEXT", "名单校验键"),
                    ("roster_id", "TEXT", "绑定名单ID"),
                    ("relation_type", "TEXT", "关系类型"),
                    ("target_label", "TEXT", "目标对象"),
                    ("submitted_date", "TEXT", "提交日期 YYYY-MM-DD"),
                    ("submitted_hour", "INTEGER", "提交小时 0-23"),
                ],
            },
            {
                "name": "identity_kv",
                "desc": "进入前采集的身份字段（键值展开）",
                "columns": [
                    ("row_id", "INTEGER", "自增行ID"),
                    ("submission_id", "TEXT", "票据ID"),
                    ("field_key", "TEXT", "字段键"),
                    ("field_label", "TEXT", "字段标签（当前等同字段键）"),
                    ("field_value", "TEXT", "字段值"),
                ],
            },
            {
                "name": "question_defs",
                "desc": "问卷题目定义（设计结构）",
                "columns": [
                    ("question_id", "TEXT", "题目ID"),
                    ("question_title", "TEXT", "题干"),
                    ("question_type", "TEXT", "题型"),
                    ("required", "INTEGER", "1=必填"),
                    ("repeat_from", "TEXT", "循环来源"),
                    ("repeat_filter", "TEXT", "循环筛选 all/self/peer"),
                    ("is_loop_question", "INTEGER", "1=循环题"),
                ],
            },
            {
                "name": "question_options",
                "desc": "题目选项定义（设计结构）",
                "columns": [
                    ("row_id", "INTEGER", "自增行ID"),
                    ("question_id", "TEXT", "题目ID"),
                    ("option_index", "INTEGER", "选项顺序"),
                    ("option_value", "TEXT", "选项值"),
                ],
            },
            {
                "name": "answers",
                "desc": "答案明细（每个原子答案一行）",
                "columns": [
                    ("row_id", "INTEGER", "自增行ID"),
                    ("submission_id", "TEXT", "票据ID"),
                    ("question_id", "TEXT", "题目ID"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("value_type", "TEXT", "值类型: number/text/list/json/null"),
                    ("value_text", "TEXT", "文本值"),
                    ("value_num", "REAL", "数值值"),
                    ("value_json", "TEXT", "原始JSON值"),
                ],
            },
            {
                "name": "answer_options",
                "desc": "选项明细（单选/多选展开）",
                "columns": [
                    ("row_id", "INTEGER", "自增行ID"),
                    ("submission_id", "TEXT", "票据ID"),
                    ("question_id", "TEXT", "题目ID"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("option_value", "TEXT", "被选中的选项值"),
                ],
            },
            {
                "name": "v_scores",
                "desc": "数值答案视图（评分/滑杆等）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("question_id", "TEXT", "题目ID"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("value_num", "REAL", "数值"),
                ],
            },
            {
                "name": "v_text_answers",
                "desc": "文本答案视图（非空文本）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("question_id", "TEXT", "题目ID"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("value_text", "TEXT", "文本值"),
                ],
            },
            {
                "name": "v_answers_enriched",
                "desc": "答案明细增强视图（已关联提交信息与题目定义）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("submitted_at", "TEXT", "提交时间"),
                    ("submitted_date", "TEXT", "提交日期"),
                    ("source", "TEXT", "来源"),
                    ("respondent_name", "TEXT", "填写者姓名"),
                    ("respondent_code", "TEXT", "填写者编号"),
                    ("question_id", "TEXT", "题目ID"),
                    ("question_title", "TEXT", "题干"),
                    ("question_type", "TEXT", "题型"),
                    ("question_repeat_from", "TEXT", "题目循环来源"),
                    ("question_repeat_filter", "TEXT", "题目循环筛选"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("value_type", "TEXT", "值类型"),
                    ("value_text", "TEXT", "文本值"),
                    ("value_num", "REAL", "数值值"),
                ],
            },
            {
                "name": "v_answer_options_enriched",
                "desc": "选项答案增强视图（已关联提交信息与题目定义）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("submitted_at", "TEXT", "提交时间"),
                    ("submitted_date", "TEXT", "提交日期"),
                    ("source", "TEXT", "来源"),
                    ("respondent_name", "TEXT", "填写者姓名"),
                    ("respondent_code", "TEXT", "填写者编号"),
                    ("question_id", "TEXT", "题目ID"),
                    ("question_title", "TEXT", "题干"),
                    ("question_type", "TEXT", "题型"),
                    ("question_repeat_from", "TEXT", "题目循环来源"),
                    ("question_repeat_filter", "TEXT", "题目循环筛选"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("option_value", "TEXT", "选项值"),
                ],
            },
            {
                "name": "v_scores_enriched",
                "desc": "数值答案增强视图（评分聚合推荐入口）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("submitted_at", "TEXT", "提交时间"),
                    ("submitted_date", "TEXT", "提交日期"),
                    ("source", "TEXT", "来源"),
                    ("respondent_name", "TEXT", "填写者姓名"),
                    ("respondent_code", "TEXT", "填写者编号"),
                    ("verified_member_key", "TEXT", "名单校验键"),
                    ("question_id", "TEXT", "题目ID"),
                    ("question_title", "TEXT", "题干"),
                    ("question_type", "TEXT", "题型"),
                    ("question_repeat_from", "TEXT", "题目循环来源"),
                    ("question_repeat_filter", "TEXT", "题目循环筛选"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("value_num", "REAL", "数值"),
                ],
            },
            {
                "name": "v_scores_identity",
                "desc": "数值答案 + 提交主表宽表（含动态身份列，便于直接筛选）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("verified_member_key", "TEXT", "名单校验键"),
                    ("verified_member_key_<auto_alias>", "TEXT", "动态身份列（如 verified_member_key_xing_ming）"),
                    ("question_id", "TEXT", "题目ID"),
                    ("question_title", "TEXT", "题干"),
                    ("question_type", "TEXT", "题型"),
                    ("question_repeat_from", "TEXT", "题目循环来源"),
                    ("question_repeat_filter", "TEXT", "题目循环筛选"),
                    ("repeat_at", "TEXT", "当前循环项值"),
                    ("value_num", "REAL", "数值"),
                ],
            },
            {
                "name": "v_identity_enriched",
                "desc": "身份字段增强视图（身份字段 + 提交信息）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("submitted_at", "TEXT", "提交时间"),
                    ("submitted_date", "TEXT", "提交日期"),
                    ("source", "TEXT", "来源"),
                    ("respondent_name", "TEXT", "填写者姓名"),
                    ("respondent_code", "TEXT", "填写者编号"),
                    ("verified_member_key", "TEXT", "名单校验键"),
                    ("field_key", "TEXT", "字段键"),
                    ("field_label", "TEXT", "字段标签"),
                    ("field_value", "TEXT", "字段值"),
                ],
            },
            {
                "name": "v_submissions_identity",
                "desc": "提交主表 + 动态身份列（按名单校验项自动生成 verified_member_key_* 列）",
                "columns": [
                    ("submission_id", "TEXT", "票据ID"),
                    ("verified_member_key", "TEXT", "名单校验键"),
                    ("verified_member_key_<auto_alias>", "TEXT", "动态身份字段值（如 verified_member_key_xing_ming）"),
                ],
            },
        ]

    def _try_float(self, value: Any) -> Optional[float]:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            t = value.strip()
            if not t:
                return None
            try:
                return float(t)
            except Exception:
                return None
        return None

    def _insert_answer_rows(
        self,
        conn: sqlite3.Connection,
        submission_id: str,
        question: Dict[str, Any],
        value: Any,
        repeat_at: str = "",
    ) -> None:
        qid = str(question.get("id", "")).strip()
        if not qid:
            return
        qtype = str(question.get("type", "")).strip()
        rep_at = str(repeat_at or "").strip()

        value_text = ""
        value_num: Optional[float] = None
        value_type = "null"
        if isinstance(value, list):
            value_text = " | ".join(str(x) for x in value)
            value_type = "list"
        elif isinstance(value, dict):
            value_text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
            value_type = "json"
        elif value is None:
            value_text = ""
            value_type = "null"
        else:
            value_text = str(value)
            value_num = self._try_float(value)
            value_type = "number" if value_num is not None else "text"
        value_json = json.dumps(value, ensure_ascii=False, separators=(",", ":"))

        conn.execute(
            """
            INSERT INTO answers (
                submission_id, question_id, repeat_at, value_type, value_text, value_num, value_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                submission_id,
                qid,
                rep_at,
                value_type,
                value_text,
                value_num,
                value_json,
            ),
        )

        if qtype == "single":
            if isinstance(value, str) and value.strip():
                conn.execute(
                    """
                    INSERT INTO answer_options (
                        submission_id, question_id, repeat_at, option_value
                    )
                    VALUES (?, ?, ?, ?)
                    """,
                    (submission_id, qid, rep_at, value.strip()),
                )
            return
        if qtype == "multi" and isinstance(value, list):
            for item in value:
                option = str(item).strip()
                if not option:
                    continue
                conn.execute(
                    """
                    INSERT INTO answer_options (
                        submission_id, question_id, repeat_at, option_value
                    )
                    VALUES (?, ?, ?, ?)
                    """,
                    (submission_id, qid, rep_at, option),
                )

    def _normalize_sql_validation_rules_from_schema(self, schema: Dict[str, Any]) -> List[Dict[str, Any]]:
        meta = schema.get("meta", {}) if isinstance(schema, dict) else {}
        if not isinstance(meta, dict):
            return []
        rules_raw = meta.get("validation_rules", [])
        if not isinstance(rules_raw, list):
            return []
        out: List[Dict[str, Any]] = []
        for idx, rule in enumerate(rules_raw, start=1):
            if not isinstance(rule, dict):
                continue
            if str(rule.get("type", "")).strip().lower() != "sql_aggregate":
                continue
            sql_text = str(rule.get("sql", "")).strip()
            if not sql_text:
                continue
            sql_normalized = self._normalize_live_rule_sql(sql_text)
            op = self._normalize_live_rule_op(str(rule.get("op", "lte")).strip())
            if op not in {"equals", "not_equals", "gt", "gte", "lt", "lte", "between", "not_between"}:
                raise ServiceError(f"联合规则 #{idx} 的比较方式无效。")
            value = self._parse_rule_numeric(rule.get("value"), f"联合规则 #{idx} 的目标值")
            value2: Optional[float] = None
            if op in {"between", "not_between"}:
                value2 = self._parse_rule_numeric(rule.get("value2"), f"联合规则 #{idx} 的区间上限值")
            out.append(
                {
                    "name": str(rule.get("name", "")).strip() or f"联合规则#{idx}",
                    "sql": sql_normalized,
                    "op": op,
                    "value": value,
                    "value2": value2,
                    "message": str(rule.get("message", "")).strip() or f"联合规则 #{idx} 未通过。",
                }
            )
        return out

    def _build_live_rule_model(
        self,
        questionnaire: Dict[str, Any],
        answers: Dict[str, Any],
        respondent_name: str,
        respondent_code: str,
        respondent_identity: Optional[Dict[str, Any]],
        current_member_key: str,
    ) -> sqlite3.Connection:
        schema = normalize_schema(questionnaire.get("schema", {}))
        questions = schema.get("questions", [])
        question_by_id: Dict[str, Dict[str, Any]] = {}
        for q in questions:
            qid = str(q.get("id", "")).strip()
            if qid:
                question_by_id[qid] = q

        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            PRAGMA temp_store = MEMORY;

            CREATE TABLE submissions (
                submission_id TEXT PRIMARY KEY,
                questionnaire_id TEXT,
                respondent_name TEXT,
                respondent_code TEXT,
                verified_member_key TEXT
            );

            CREATE TABLE identity_kv (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id TEXT NOT NULL,
                field_key TEXT NOT NULL,
                field_label TEXT,
                field_value TEXT
            );

            CREATE TABLE question_defs (
                question_id TEXT PRIMARY KEY,
                question_title TEXT,
                question_type TEXT,
                repeat_from TEXT,
                repeat_filter TEXT
            );

            CREATE TABLE question_options (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                question_id TEXT NOT NULL,
                option_index INTEGER NOT NULL,
                option_value TEXT
            );

            CREATE TABLE answers (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id TEXT NOT NULL,
                question_id TEXT NOT NULL,
                repeat_at TEXT,
                value_type TEXT,
                value_text TEXT,
                value_num REAL,
                value_json TEXT
            );

            CREATE TABLE answer_options (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id TEXT NOT NULL,
                question_id TEXT NOT NULL,
                repeat_at TEXT,
                option_value TEXT
            );

            CREATE VIEW v_scores AS
                SELECT submission_id, question_id, repeat_at, value_num
                FROM answers
                WHERE value_num IS NOT NULL;

            CREATE VIEW v_text_answers AS
                SELECT submission_id, question_id, repeat_at, value_text
                FROM answers
                WHERE COALESCE(TRIM(value_text), '') <> '';

            CREATE VIEW v_answers_enriched AS
                SELECT
                    a.submission_id,
                    s.respondent_name,
                    s.respondent_code,
                    s.verified_member_key,
                    a.question_id,
                    q.question_title AS question_title,
                    q.question_type AS question_type,
                    q.repeat_from AS question_repeat_from,
                    q.repeat_filter AS question_repeat_filter,
                    a.repeat_at,
                    a.value_type,
                    a.value_text,
                    a.value_num,
                    a.value_json
                FROM answers a
                LEFT JOIN submissions s ON s.submission_id = a.submission_id
                LEFT JOIN question_defs q ON q.question_id = a.question_id;

            CREATE VIEW v_answer_options_enriched AS
                SELECT
                    o.submission_id,
                    s.respondent_name,
                    s.respondent_code,
                    s.verified_member_key,
                    o.question_id,
                    q.question_title AS question_title,
                    q.question_type AS question_type,
                    q.repeat_from AS question_repeat_from,
                    q.repeat_filter AS question_repeat_filter,
                    o.repeat_at,
                    o.option_value
                FROM answer_options o
                LEFT JOIN submissions s ON s.submission_id = o.submission_id
                LEFT JOIN question_defs q ON q.question_id = o.question_id;

            CREATE VIEW v_scores_enriched AS
                SELECT
                    a.submission_id,
                    s.respondent_name,
                    s.respondent_code,
                    s.verified_member_key,
                    a.question_id,
                    q.question_title AS question_title,
                    q.question_type AS question_type,
                    q.repeat_from AS question_repeat_from,
                    q.repeat_filter AS question_repeat_filter,
                    a.repeat_at,
                    a.value_num
                FROM answers a
                LEFT JOIN submissions s ON s.submission_id = a.submission_id
                LEFT JOIN question_defs q ON q.question_id = a.question_id
                WHERE a.value_num IS NOT NULL;

            CREATE VIEW v_identity_enriched AS
                SELECT
                    k.submission_id,
                    s.respondent_name,
                    s.respondent_code,
                    s.verified_member_key,
                    k.field_key,
                    k.field_label,
                    k.field_value
                FROM identity_kv k
                LEFT JOIN submissions s ON s.submission_id = k.submission_id;
            """
        )

        collect_fields = self._questionnaire_collect_fields(questionnaire)
        field_defs: List[Dict[str, str]] = []
        seen_field_keys: set[str] = set()
        for item in collect_fields:
            key = str(item.get("key", "")).strip()
            if not key or key in seen_field_keys:
                continue
            seen_field_keys.add(key)
            field_defs.append({"key": key, "label": str(item.get("label", "")).strip() or key})
        roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
        if roster_id:
            for col in self.get_roster_columns(roster_id):
                key = str(col.get("key", "")).strip()
                if not key or key in seen_field_keys:
                    continue
                seen_field_keys.add(key)
                field_defs.append({"key": key, "label": str(col.get("label", "")).strip() or key})
        alias_map = self._identity_alias_map(field_defs)
        for key in [str(item.get("key", "")).strip() for item in field_defs]:
            alias = alias_map.get(key, "")
            if not alias:
                continue
            col_name = f"verified_member_key_{alias}"
            conn.execute(f"ALTER TABLE submissions ADD COLUMN {self._sql_quote_ident(col_name)} TEXT")

        current_submission_id = "__current__"
        conn.execute(
            """
            INSERT INTO submissions (
                submission_id, questionnaire_id, respondent_name, respondent_code, verified_member_key
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                current_submission_id,
                str(questionnaire.get("id", "")).strip(),
                str(respondent_name or "").strip(),
                str(respondent_code or "").strip(),
                str(current_member_key or "").strip(),
            ),
        )
        identity = respondent_identity if isinstance(respondent_identity, dict) else {}
        dynamic_updates: Dict[str, str] = {}
        for field in field_defs:
            key = str(field.get("key", "")).strip()
            alias = alias_map.get(key, "")
            if not key or not alias:
                continue
            value = str(identity.get(key, "")).strip()
            if not value and key == "member_name":
                value = str(respondent_name or "").strip()
            if not value and key == "member_code":
                value = str(respondent_code or "").strip()
            if value:
                dynamic_updates[f"verified_member_key_{alias}"] = value
        if dynamic_updates:
            set_sql = ", ".join([f"{self._sql_quote_ident(col)} = ?" for col in dynamic_updates.keys()])
            params = list(dynamic_updates.values()) + [current_submission_id]
            conn.execute(f"UPDATE submissions SET {set_sql} WHERE submission_id = ?", tuple(params))

        identity_keys_seen: set[str] = set()
        for field in field_defs:
            key = str(field.get("key", "")).strip()
            if not key:
                continue
            label = str(field.get("label", "")).strip() or key
            value = str(identity.get(key, "")).strip()
            if not value and key == "member_name":
                value = str(respondent_name or "").strip()
            if not value and key == "member_code":
                value = str(respondent_code or "").strip()
            identity_keys_seen.add(key)
            conn.execute(
                """
                INSERT INTO identity_kv (submission_id, field_key, field_label, field_value)
                VALUES (?, ?, ?, ?)
                """,
                (current_submission_id, key, label, value),
            )
        for key, raw_value in identity.items():
            field_key = str(key).strip()
            if not field_key or field_key in identity_keys_seen:
                continue
            field_value = str(raw_value).strip()
            identity_keys_seen.add(field_key)
            conn.execute(
                """
                INSERT INTO identity_kv (submission_id, field_key, field_label, field_value)
                VALUES (?, ?, ?, ?)
                """,
                (current_submission_id, field_key, field_key, field_value),
            )

        schema_meta = schema.get("meta", {}) if isinstance(schema.get("meta", {}), dict) else {}
        for q in questions:
            conn.execute(
                """
                INSERT INTO question_defs (question_id, question_title, question_type, repeat_from, repeat_filter)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    str(q.get("id", "")).strip(),
                    str(q.get("title", "")).strip(),
                    str(q.get("type", "")).strip(),
                    str(q.get("repeat_from", "")).strip(),
                    str(q.get("repeat_filter", "all")).strip(),
                ),
            )
            options = q.get("options", [])
            if isinstance(options, list):
                for idx, option in enumerate(options, start=1):
                    opt = str(option).strip()
                    if not opt:
                        continue
                    conn.execute(
                        """
                        INSERT INTO question_options (question_id, option_index, option_value)
                        VALUES (?, ?, ?)
                        """,
                        (str(q.get("id", "")).strip(), idx, opt),
                    )

        base_columns = [
            "submission_id",
            "questionnaire_id",
            "respondent_name",
            "respondent_code",
            "verified_member_key",
        ]
        alias_map_all = self._identity_alias_map(field_defs, extra_keys=sorted(identity_keys_seen))
        select_parts = [f"s.{col}" for col in base_columns]
        for key in sorted(identity_keys_seen):
            alias = alias_map_all.get(key, "")
            if not alias:
                continue
            col_name = f"verified_member_key_{alias}"
            expr = (
                f"MAX(CASE WHEN k.field_key={self._sql_quote_literal(key)} THEN k.field_value END) "
                f"AS {self._sql_quote_ident(col_name)}"
            )
            select_parts.append(expr)
        group_by_sql = ", ".join([f"s.{col}" for col in base_columns])
        view_sql = (
            "CREATE VIEW v_submissions_identity AS\n"
            "SELECT\n  "
            + ",\n  ".join(select_parts)
            + "\nFROM submissions s\n"
            "LEFT JOIN identity_kv k ON k.submission_id = s.submission_id\n"
            f"GROUP BY {group_by_sql};"
        )
        conn.execute("DROP VIEW IF EXISTS v_submissions_identity")
        conn.execute(view_sql)
        conn.execute("DROP VIEW IF EXISTS v_scores_identity")
        conn.execute(
            """
            CREATE VIEW v_scores_identity AS
            SELECT
                si.*,
                a.question_id,
                q.question_title AS question_title,
                q.question_type AS question_type,
                q.repeat_from AS question_repeat_from,
                q.repeat_filter AS question_repeat_filter,
                a.repeat_at,
                a.value_num
            FROM v_submissions_identity si
            JOIN answers a ON a.submission_id = si.submission_id
            LEFT JOIN question_defs q ON q.question_id = a.question_id
            WHERE a.value_num IS NOT NULL;
            """
        )

        for qid, raw_value in answers.items():
            qid_s = str(qid).strip()
            if not qid_s:
                continue
            question = question_by_id.get(qid_s, {"id": qid_s, "title": qid_s, "type": "text"})
            if isinstance(raw_value, dict):
                for repeat_key, repeat_value in raw_value.items():
                    rk = str(repeat_key).strip()
                    if not rk:
                        continue
                    repeat_value_at = self._resolve_repeat_item_value(question, rk, schema_meta=schema_meta)
                    self._insert_answer_rows(
                        conn=conn,
                        submission_id=current_submission_id,
                        question=question,
                        value=repeat_value,
                        repeat_at=repeat_value_at,
                    )
            else:
                self._insert_answer_rows(
                    conn=conn,
                    submission_id=current_submission_id,
                    question=question,
                    value=raw_value,
                )
        conn.commit()
        return conn

    def _evaluate_sql_validation_rules(
        self,
        questionnaire: Dict[str, Any],
        answers: Dict[str, Any],
        respondent_name: str,
        respondent_code: str,
        respondent_identity: Optional[Dict[str, Any]],
        current_member_key: str,
    ) -> List[str]:
        schema = questionnaire.get("schema", {}) if isinstance(questionnaire.get("schema", {}), dict) else {}
        rules = self._normalize_sql_validation_rules_from_schema(schema)
        if not rules:
            return []
        conn = self._build_live_rule_model(
            questionnaire=questionnaire,
            answers=answers,
            respondent_name=respondent_name,
            respondent_code=respondent_code,
            respondent_identity=respondent_identity,
            current_member_key=current_member_key,
        )
        errors: List[str] = []
        try:
            for idx, rule in enumerate(rules, start=1):
                sql_text = self._compose_live_rule_sql(str(rule.get("sql", "")).strip(), submission_id="__current__")
                try:
                    row = conn.execute(sql_text).fetchone()
                except sqlite3.Error as exc:
                    raise ServiceError(f"联合规则 #{idx} SQL 执行失败: {exc}") from exc
                value_raw: Any = None
                if row is not None and len(row.keys()) >= 1:
                    value_raw = row[0]
                if value_raw is None:
                    actual = 0.0
                else:
                    actual = self._try_float(value_raw)
                    if actual is None:
                        raise ServiceError(
                            f"联合规则 #{idx} 的 SQL 返回值不是数字（得到：{value_raw}）。"
                        )
                op = str(rule.get("op", "lte")).strip().lower()
                target = float(rule.get("value", 0))
                target2_raw = rule.get("value2")
                target2 = float(target2_raw) if target2_raw is not None else None
                if not self._live_rule_compare(actual, op, target, target2):
                    errors.append(str(rule.get("message", "")).strip() or f"联合规则 #{idx} 未通过。")
        finally:
            conn.close()
        return errors

    def check_live_rules(
        self,
        questionnaire_id: str,
        answers: Dict[str, Any],
        respondent_name: str = "",
        respondent_code: str = "",
        respondent_identity: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        qid = str(questionnaire_id or "").strip()
        questionnaire = self.get_questionnaire(qid)
        if not questionnaire:
            raise ServiceError("问卷不存在。")
        schema = questionnaire.get("schema", {}) if isinstance(questionnaire.get("schema", {}), dict) else {}
        raw_identity = respondent_identity if isinstance(respondent_identity, dict) else {}
        if respondent_name.strip():
            raw_identity["member_name"] = respondent_name.strip()
        if respondent_code.strip():
            raw_identity["member_code"] = respondent_code.strip()
        current_member_key = self._resolve_current_member_key(
            questionnaire=questionnaire,
            verified_member_key="",
            respondent_code=str(raw_identity.get("member_code", "")).strip(),
            respondent_name=str(raw_identity.get("member_name", "")).strip(),
            respondent_identity=raw_identity,
        )
        answers_for_validate = dict(answers if isinstance(answers, dict) else {})
        self._inject_schema_list_sources(
            schema=schema,
            answers_for_validate=answers_for_validate,
            current_member_key=current_member_key,
        )
        self._inject_legacy_repeat_source(schema=schema, answers_for_validate=answers_for_validate)
        ok, validate_errors, cleaned_answers = validate_answers(schema, answers_for_validate)
        hint_errors: List[str] = []
        if not ok:
            for err in validate_errors:
                text = str(err).strip()
                if not text:
                    continue
                if "为必填项" in text or "循环项中存在未填写内容" in text:
                    continue
                hint_errors.append(text)
        rule_errors = self._evaluate_sql_validation_rules(
            questionnaire=questionnaire,
            answers=cleaned_answers,
            respondent_name=str(raw_identity.get("member_name", "")).strip(),
            respondent_code=str(raw_identity.get("member_code", "")).strip(),
            respondent_identity=raw_identity,
            current_member_key=current_member_key,
        )
        return {
            "pass": len(rule_errors) == 0,
            "rule_errors": rule_errors,
            "hints": hint_errors[:6],
            "auto_filter_suffix": self.live_rule_auto_filter_suffix(),
        }

    def _build_query_model(self, questionnaire_id: str) -> sqlite3.Connection:
        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            raise ServiceError("问卷不存在。")
        collect_fields = self._questionnaire_collect_fields(questionnaire)
        field_defs: List[Dict[str, str]] = []
        seen_field_keys: set[str] = set()
        for item in collect_fields:
            key = str(item.get("key", "")).strip()
            if not key or key in seen_field_keys:
                continue
            seen_field_keys.add(key)
            field_defs.append({"key": key, "label": str(item.get("label", "")).strip() or key})
        roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
        if roster_id:
            for col in self.get_roster_columns(roster_id):
                key = str(col.get("key", "")).strip()
                if not key or key in seen_field_keys:
                    continue
                seen_field_keys.add(key)
                field_defs.append({"key": key, "label": str(col.get("label", "")).strip() or key})
        collect_keys = [str(item.get("key", "")).strip() for item in field_defs if str(item.get("key", "")).strip()]
        alias_map_base = self._identity_alias_map(field_defs)
        payloads = self.decrypt_submission_payloads(questionnaire_id)
        submissions_meta = {
            str(row.get("id", "")).strip(): row for row in self.db.list_submissions(questionnaire_id=questionnaire_id)
        }
        schema = normalize_schema(questionnaire.get("schema", {}))
        questions = schema.get("questions", [])
        question_by_id: Dict[str, Dict[str, Any]] = {}
        for q in questions:
            qid = str(q.get("id", "")).strip()
            if not qid:
                continue
            question_by_id[qid] = q

        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            PRAGMA temp_store = MEMORY;

            CREATE TABLE submissions (
                submission_id TEXT PRIMARY KEY,
                questionnaire_id TEXT,
                questionnaire_version INTEGER,
                submitted_at TEXT,
                source TEXT,
                anonymous INTEGER,
                respondent_name TEXT,
                respondent_code TEXT,
                verified_member_key TEXT,
                roster_id TEXT,
                relation_type TEXT,
                target_label TEXT,
                submitted_date TEXT,
                submitted_hour INTEGER
            );

            CREATE TABLE identity_kv (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id TEXT NOT NULL,
                field_key TEXT NOT NULL,
                field_label TEXT,
                field_value TEXT
            );

            CREATE TABLE question_defs (
                question_id TEXT PRIMARY KEY,
                question_title TEXT,
                question_type TEXT,
                required INTEGER,
                repeat_from TEXT,
                repeat_filter TEXT,
                is_loop_question INTEGER
            );

            CREATE TABLE question_options (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                question_id TEXT NOT NULL,
                option_index INTEGER NOT NULL,
                option_value TEXT
            );

            CREATE TABLE answers (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id TEXT NOT NULL,
                question_id TEXT NOT NULL,
                repeat_at TEXT,
                value_type TEXT,
                value_text TEXT,
                value_num REAL,
                value_json TEXT
            );

            CREATE TABLE answer_options (
                row_id INTEGER PRIMARY KEY AUTOINCREMENT,
                submission_id TEXT NOT NULL,
                question_id TEXT NOT NULL,
                repeat_at TEXT,
                option_value TEXT
            );

            CREATE INDEX idx_answers_sid ON answers(submission_id);
            CREATE INDEX idx_answers_qid ON answers(question_id);
            CREATE INDEX idx_answer_opts_sid ON answer_options(submission_id);
            CREATE INDEX idx_answer_opts_qid ON answer_options(question_id);
            CREATE INDEX idx_identity_sid ON identity_kv(submission_id);
            CREATE INDEX idx_identity_key ON identity_kv(field_key);

            CREATE VIEW v_scores AS
                SELECT submission_id, question_id, repeat_at, value_num
                FROM answers
                WHERE value_num IS NOT NULL;

            CREATE VIEW v_text_answers AS
                SELECT submission_id, question_id, repeat_at, value_text
                FROM answers
                WHERE COALESCE(TRIM(value_text), '') <> '';

            CREATE VIEW v_answers_enriched AS
                SELECT
                    a.submission_id,
                    s.submitted_at,
                    s.submitted_date,
                    s.source,
                    s.respondent_name,
                    s.respondent_code,
                    a.question_id,
                    q.question_title AS question_title,
                    q.question_type AS question_type,
                    q.repeat_from AS question_repeat_from,
                    q.repeat_filter AS question_repeat_filter,
                    a.repeat_at,
                    a.value_type,
                    a.value_text,
                    a.value_num,
                    a.value_json
                FROM answers a
                LEFT JOIN submissions s ON s.submission_id = a.submission_id
                LEFT JOIN question_defs q ON q.question_id = a.question_id;

            CREATE VIEW v_answer_options_enriched AS
                SELECT
                    o.submission_id,
                    s.submitted_at,
                    s.submitted_date,
                    s.source,
                    s.respondent_name,
                    s.respondent_code,
                    o.question_id,
                    q.question_title AS question_title,
                    q.question_type AS question_type,
                    q.repeat_from AS question_repeat_from,
                    q.repeat_filter AS question_repeat_filter,
                    o.repeat_at,
                    o.option_value
                FROM answer_options o
                LEFT JOIN submissions s ON s.submission_id = o.submission_id
                LEFT JOIN question_defs q ON q.question_id = o.question_id;

            CREATE VIEW v_scores_enriched AS
                SELECT
                    a.submission_id,
                    s.submitted_at,
                    s.submitted_date,
                    s.source,
                    s.respondent_name,
                    s.respondent_code,
                    s.verified_member_key,
                    a.question_id,
                    q.question_title AS question_title,
                    q.question_type AS question_type,
                    q.repeat_from AS question_repeat_from,
                    q.repeat_filter AS question_repeat_filter,
                    a.repeat_at,
                    a.value_num
                FROM answers a
                LEFT JOIN submissions s ON s.submission_id = a.submission_id
                LEFT JOIN question_defs q ON q.question_id = a.question_id
                WHERE a.value_num IS NOT NULL;

            CREATE VIEW v_identity_enriched AS
                SELECT
                    k.submission_id,
                    s.submitted_at,
                    s.submitted_date,
                    s.source,
                    s.respondent_name,
                    s.respondent_code,
                    s.verified_member_key,
                    k.field_key,
                    k.field_label,
                    k.field_value
                FROM identity_kv k
                LEFT JOIN submissions s ON s.submission_id = k.submission_id;
            """
        )

        for key in collect_keys:
            alias = alias_map_base.get(key, "")
            if not alias:
                continue
            col_name = f"verified_member_key_{alias}"
            conn.execute(f"ALTER TABLE submissions ADD COLUMN {self._sql_quote_ident(col_name)} TEXT")

        identity_keys_seen: set[str] = set(collect_keys)
        schema_meta = schema.get("meta", {}) if isinstance(schema.get("meta", {}), dict) else {}
        for q in questions:
            conn.execute(
                """
                INSERT INTO question_defs (
                    question_id, question_title, question_type, required, repeat_from, repeat_filter, is_loop_question
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(q.get("id", "")).strip(),
                    str(q.get("title", "")).strip(),
                    str(q.get("type", "")).strip(),
                    1 if bool(q.get("required", False)) else 0,
                    str(q.get("repeat_from", "")).strip(),
                    str(q.get("repeat_filter", "all")).strip(),
                    1 if str(q.get("repeat_from", "")).strip() else 0,
                ),
            )
            options = q.get("options", [])
            if isinstance(options, list):
                for idx, option in enumerate(options, start=1):
                    opt = str(option).strip()
                    if not opt:
                        continue
                    conn.execute(
                        """
                        INSERT INTO question_options (question_id, option_index, option_value)
                        VALUES (?, ?, ?)
                        """,
                        (str(q.get("id", "")).strip(), idx, opt),
                    )

        for payload in payloads:
            submission_id = str(payload.get("submission_id", "")).strip()
            if not submission_id:
                continue
            meta = submissions_meta.get(submission_id, {})
            respondent = payload.get("respondent", {}) if isinstance(payload.get("respondent", {}), dict) else {}
            verified = payload.get("verified", {}) if isinstance(payload.get("verified", {}), dict) else {}
            context = payload.get("context", {}) if isinstance(payload.get("context", {}), dict) else {}

            source = str(payload.get("submitted_from", "")).strip() or str(meta.get("source", "")).strip()
            submitted_at = str(meta.get("submitted_at", "")).strip()
            anonymous = bool(respondent.get("anonymous", meta.get("anonymous", False)))
            respondent_name = str(respondent.get("name", "")).strip() or str(meta.get("respondent_name", "")).strip()
            respondent_code = str(respondent.get("code", "")).strip() or str(meta.get("respondent_code", "")).strip()
            verified_member_key = str(verified.get("member_key", "")).strip() or str(meta.get("verified_member_key", "")).strip()
            roster_id = str(verified.get("roster_id", "")).strip() or str(meta.get("roster_id", "")).strip()
            relation_type = str(context.get("relation_type", "")).strip() or str(meta.get("session_label", "")).strip()
            target_label = str(context.get("target_label", "")).strip() or str(meta.get("target_label", "")).strip()
            submitted_date = ""
            submitted_hour: Optional[int] = None
            if submitted_at:
                try:
                    dt = parse_iso(submitted_at)
                    submitted_date = dt.date().isoformat()
                    submitted_hour = int(dt.hour)
                except Exception:
                    submitted_date = submitted_at[:10]
                    submitted_hour = None

            conn.execute(
                """
                INSERT INTO submissions (
                    submission_id, questionnaire_id, questionnaire_version, submitted_at,
                    source, anonymous, respondent_name, respondent_code, verified_member_key,
                    roster_id, relation_type, target_label, submitted_date, submitted_hour
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    submission_id,
                    questionnaire_id,
                    int(payload.get("questionnaire_version", meta.get("questionnaire_version", 1)) or 1),
                    submitted_at,
                    source,
                    1 if anonymous else 0,
                    respondent_name,
                    respondent_code,
                    verified_member_key,
                    roster_id,
                    relation_type,
                    target_label,
                    submitted_date,
                    submitted_hour,
                ),
            )
            identity_data = respondent.get("identity_data", {}) if isinstance(respondent.get("identity_data", {}), dict) else {}

            dynamic_updates: Dict[str, str] = {}
            for key in collect_keys:
                alias = alias_map_base.get(key, "")
                if not alias:
                    continue
                value = str(identity_data.get(key, "")).strip()
                if not value and key == "member_name":
                    value = respondent_name
                if not value and key == "member_code":
                    value = respondent_code
                if not value:
                    continue
                dynamic_updates[f"verified_member_key_{alias}"] = value
            if dynamic_updates:
                set_sql = ", ".join([f"{self._sql_quote_ident(col)} = ?" for col in dynamic_updates.keys()])
                params = list(dynamic_updates.values()) + [submission_id]
                conn.execute(
                    f"UPDATE submissions SET {set_sql} WHERE submission_id = ?",
                    tuple(params),
                )

            for key, raw_value in identity_data.items():
                field_key = str(key).strip()
                field_value = str(raw_value).strip()
                if not field_key:
                    continue
                identity_keys_seen.add(field_key)
                conn.execute(
                    """
                    INSERT INTO identity_kv (submission_id, field_key, field_label, field_value)
                    VALUES (?, ?, ?, ?)
                    """,
                    (submission_id, field_key, field_key, field_value),
                )

            answers = payload.get("answers", {})
            if not isinstance(answers, dict):
                continue
            for qid, raw in answers.items():
                qid_s = str(qid).strip()
                if not qid_s:
                    continue
                question = question_by_id.get(qid_s, {"id": qid_s, "title": qid_s, "type": "text"})
                if isinstance(raw, dict):
                    for repeat_key, repeat_value in raw.items():
                        rk = str(repeat_key).strip()
                        if not rk:
                            continue
                        repeat_value_at = self._resolve_repeat_item_value(question, rk, schema_meta=schema_meta)
                        self._insert_answer_rows(
                            conn=conn,
                            submission_id=submission_id,
                            question=question,
                            value=repeat_value,
                            repeat_at=repeat_value_at,
                        )
                else:
                    self._insert_answer_rows(
                        conn=conn,
                        submission_id=submission_id,
                        question=question,
                        value=raw,
                    )

        alias_map = self._identity_alias_map(field_defs, extra_keys=sorted(identity_keys_seen))
        base_columns = [
            "submission_id",
            "questionnaire_id",
            "questionnaire_version",
            "submitted_at",
            "source",
            "anonymous",
            "respondent_name",
            "respondent_code",
            "verified_member_key",
            "roster_id",
            "relation_type",
            "target_label",
            "submitted_date",
            "submitted_hour",
        ]
        select_parts = [f"s.{col}" for col in base_columns]
        for key in sorted(identity_keys_seen):
            alias = alias_map.get(key, "")
            if not alias:
                continue
            col_name = f"verified_member_key_{alias}"
            expr = (
                f"MAX(CASE WHEN k.field_key={self._sql_quote_literal(key)} THEN k.field_value END) "
                f"AS {self._sql_quote_ident(col_name)}"
            )
            select_parts.append(expr)
        group_by_sql = ", ".join([f"s.{col}" for col in base_columns])
        view_sql = (
            "CREATE VIEW v_submissions_identity AS\n"
            "SELECT\n  "
            + ",\n  ".join(select_parts)
            + "\nFROM submissions s\n"
            "LEFT JOIN identity_kv k ON k.submission_id = s.submission_id\n"
            f"GROUP BY {group_by_sql};"
        )
        conn.execute("DROP VIEW IF EXISTS v_submissions_identity")
        conn.execute(view_sql)
        conn.execute("DROP VIEW IF EXISTS v_scores_identity")
        conn.execute(
            """
            CREATE VIEW v_scores_identity AS
            SELECT
                si.*,
                a.question_id,
                q.question_title AS question_title,
                q.question_type AS question_type,
                q.repeat_from AS question_repeat_from,
                q.repeat_filter AS question_repeat_filter,
                a.repeat_at,
                a.value_num
            FROM v_submissions_identity si
            JOIN answers a ON a.submission_id = si.submission_id
            LEFT JOIN question_defs q ON q.question_id = a.question_id
            WHERE a.value_num IS NOT NULL;
            """
        )

        conn.commit()
        return conn

    def query_model_schema(self, questionnaire_id: str) -> Dict[str, Any]:
        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            raise ServiceError("问卷不存在。")
        submissions = self.db.list_submissions(questionnaire_id=questionnaire_id)
        collect_fields = self._questionnaire_collect_fields(questionnaire)
        field_defs: List[Dict[str, str]] = []
        seen_field_keys: set[str] = set()
        for item in collect_fields:
            key = str(item.get("key", "")).strip()
            if not key or key in seen_field_keys:
                continue
            seen_field_keys.add(key)
            field_defs.append({"key": key, "label": str(item.get("label", "")).strip() or key})
        roster_id = str(questionnaire.get("auth_roster_id", "")).strip()
        if roster_id:
            for col in self.get_roster_columns(roster_id):
                key = str(col.get("key", "")).strip()
                if not key or key in seen_field_keys:
                    continue
                seen_field_keys.add(key)
                field_defs.append({"key": key, "label": str(col.get("label", "")).strip() or key})
        alias_map = self._identity_alias_map(field_defs)
        dynamic_identity_columns = [
            {
                "field_key": str(item.get("key", "")).strip(),
                "field_label": str(item.get("label", "")).strip() or str(item.get("key", "")).strip(),
                "column_name": f"verified_member_key_{alias_map.get(str(item.get('key', '')).strip(), 'field')}",
            }
            for item in field_defs
            if str(item.get("key", "")).strip()
        ]
        examples: List[str] = [
            "SELECT * FROM submissions ORDER BY submitted_at DESC LIMIT 50",
            "SELECT * FROM identity_kv WHERE field_key='member_code' ORDER BY field_value",
            "SELECT * FROM v_identity_enriched "
            "WHERE field_key='member_code' AND field_value LIKE '2024%' ORDER BY field_value",
            "SELECT question_id, question_title, AVG(value_num) AS avg_score "
            "FROM v_answers_enriched WHERE question_type IN ('rating','slider') "
            "GROUP BY question_id, question_title ORDER BY avg_score DESC",
            "SELECT question_id, repeat_at, AVG(value_num) AS avg_score "
            "FROM v_answers_enriched WHERE question_type='rating' "
            "GROUP BY question_id, repeat_at ORDER BY question_id, repeat_at",
            "SELECT question_id, option_value, COUNT(*) AS cnt "
            "FROM answer_options GROUP BY question_id, option_value "
            "ORDER BY question_id, cnt DESC",
            "SELECT question_id, AVG(value_num) AS avg_score FROM v_answers_enriched "
            "GROUP BY question_id ORDER BY avg_score DESC",
            "SELECT question_id, repeat_at, AVG(value_num) AS avg_score "
            "FROM v_scores WHERE repeat_at<>'' GROUP BY question_id, repeat_at "
            "HAVING COUNT(*)>=3 ORDER BY question_id, avg_score DESC",
            "SELECT submitted_date, COUNT(*) AS cnt FROM submissions "
            "GROUP BY submitted_date ORDER BY submitted_date;",
            "SELECT question_id, AVG(value_num) AS avg_score FROM v_scores GROUP BY question_id;",
            "SELECT * FROM submissions WHERE submitted_date BETWEEN '2026-01-01' AND '2026-12-31' "
            "ORDER BY submitted_at;",
            "SELECT question_id, option_value, COUNT(*) AS cnt FROM v_answer_options_enriched "
            "GROUP BY question_id, option_value ORDER BY cnt DESC;",
            "SELECT question_id, repeat_at, AVG(value_num) AS avg_score "
            "FROM v_scores_enriched WHERE question_id='q_score' "
            "GROUP BY question_id, repeat_at ORDER BY repeat_at;",
            "SELECT a.repeat_at, AVG(a.value_num) AS peer_avg "
            "FROM v_scores_enriched a JOIN v_submissions_identity s ON s.submission_id=a.submission_id "
            "WHERE a.question_id='q_score' AND a.repeat_at<>s.verified_member_key_xing_ming "
            "GROUP BY a.repeat_at ORDER BY peer_avg DESC;",
            "SELECT question_id, AVG(value_num) AS avg_score FROM v_scores "
            "WHERE question_id IN ('q_score','q_total') GROUP BY question_id HAVING AVG(value_num)>=3;",
            "SELECT s.submission_id, s.respondent_name FROM submissions s "
            "WHERE EXISTS (SELECT 1 FROM answers a WHERE a.submission_id=s.submission_id AND a.question_id='q_score');",
            "SELECT question_id, AVG(value_num) FROM v_scores "
            "WHERE value_num BETWEEN 1 AND 4 GROUP BY question_id;",
        ]
        if dynamic_identity_columns:
            first_col = dynamic_identity_columns[0].get("column_name", "")
            if first_col:
                examples.append(
                    f"SELECT submission_id, verified_member_key, {first_col} "
                    "FROM v_submissions_identity ORDER BY submitted_at DESC LIMIT 50"
                )
                examples.append(
                    f"SELECT question_id, repeat_at, AVG(value_num) AS avg_score "
                    f"FROM v_scores_identity WHERE {first_col}<>'' "
                    "GROUP BY question_id, repeat_at ORDER BY question_id, avg_score DESC"
                )
        table_defs = self._query_model_tables()
        for table in table_defs:
            if str(table.get("name", "")).strip() != "v_submissions_identity":
                continue
            cols = table.get("columns", [])
            if not isinstance(cols, list):
                cols = []
            for item in dynamic_identity_columns:
                if not isinstance(item, dict):
                    continue
                column_name = str(item.get("column_name", "")).strip()
                field_label = str(item.get("field_label", "")).strip()
                field_key = str(item.get("field_key", "")).strip()
                if not column_name:
                    continue
                cols.append((column_name, "TEXT", f"动态身份列：{field_label or field_key}"))
            table["columns"] = cols
            break
        for table in table_defs:
            if str(table.get("name", "")).strip() != "submissions":
                continue
            cols = table.get("columns", [])
            if not isinstance(cols, list):
                cols = []
            for item in dynamic_identity_columns:
                if not isinstance(item, dict):
                    continue
                column_name = str(item.get("column_name", "")).strip()
                field_label = str(item.get("field_label", "")).strip()
                field_key = str(item.get("field_key", "")).strip()
                if not column_name:
                    continue
                cols.append((column_name, "TEXT", f"动态身份列：{field_label or field_key}"))
            table["columns"] = cols
            break
        for table in table_defs:
            if str(table.get("name", "")).strip() != "v_scores_identity":
                continue
            cols = table.get("columns", [])
            if not isinstance(cols, list):
                cols = []
            for item in dynamic_identity_columns:
                if not isinstance(item, dict):
                    continue
                column_name = str(item.get("column_name", "")).strip()
                field_label = str(item.get("field_label", "")).strip()
                field_key = str(item.get("field_key", "")).strip()
                if not column_name:
                    continue
                cols.append((column_name, "TEXT", f"动态身份列：{field_label or field_key}"))
            table["columns"] = cols
            break
        live_rule_table_set = set(self._live_rule_allowed_tables())
        live_rule_table_defs: List[Dict[str, Any]] = []
        for table in table_defs:
            if not isinstance(table, dict):
                continue
            table_name = str(table.get("name", "")).strip()
            if table_name not in live_rule_table_set:
                continue
            live_rule_table_defs.append(
                {
                    "name": table_name,
                    "desc": str(table.get("desc", "")).strip(),
                    "columns": list(table.get("columns", []) if isinstance(table.get("columns", []), list) else []),
                }
            )
        live_rule_examples: List[str] = [
            "SELECT AVG(value_num) FROM v_scores_enriched WHERE question_id='q_score'",
            "SELECT COUNT(*) FROM answers a JOIN submissions s ON a.submission_id=s.submission_id "
            "WHERE a.question_id='q_score' AND a.value_num>=4 AND a.repeat_at<>s.verified_member_key_xing_ming",
            "SELECT COUNT(*) FROM v_scores a JOIN submissions s ON s.submission_id=a.submission_id "
            "WHERE a.question_id='q_score' AND a.value_num>=4 AND a.repeat_at<>s.verified_member_key_xing_ming",
            "SELECT COUNT(*) FROM answers a WHERE a.question_id='q_score' "
            "AND EXISTS (SELECT 1 FROM submissions s WHERE s.submission_id=a.submission_id "
            "AND s.verified_member_key_xing_ming<>a.repeat_at)",
            "SELECT AVG(value_num) FROM v_scores WHERE question_id='q_score' AND value_num BETWEEN 1 AND 4",
            "SELECT COUNT(*) FROM (SELECT repeat_at FROM v_scores WHERE question_id='q_score' GROUP BY repeat_at "
            "HAVING AVG(value_num)>=3) t",
        ]
        return {
            "questionnaire_id": questionnaire_id,
            "questionnaire_title": questionnaire.get("title", ""),
            "table_defs": table_defs,
            "live_rule_table_defs": live_rule_table_defs,
            "submission_count": len(submissions),
            "identity_dynamic_columns": dynamic_identity_columns,
            "examples": examples,
            "live_rule_examples": live_rule_examples,
            "live_rule_suffix": self.live_rule_auto_filter_suffix(),
        }

    def execute_sql_query(
        self,
        questionnaire_id: str,
        sql_text: str,
        row_limit: int = 5000,
    ) -> Dict[str, Any]:
        qid = str(questionnaire_id or "").strip()
        if not qid:
            raise ServiceError("问卷不存在。")
        script = self._normalize_query_script(sql_text)
        statements = self._split_sql_statements(script)
        limit = max(1, min(int(row_limit), 50000))
        conn = self._build_query_model(qid)
        try:
            result_sets: List[Dict[str, Any]] = []
            for idx, stmt in enumerate(statements, start=1):
                sql = self._normalize_select_statement(stmt)
                cursor = conn.execute(sql)
                if cursor.description is None:
                    raise ServiceError(f"第 {idx} 条语句无结果列，请使用 SELECT 查询。")
                columns = [str(col[0]) for col in cursor.description]
                fetched = cursor.fetchmany(limit + 1)
                truncated = len(fetched) > limit
                if truncated:
                    fetched = fetched[:limit]
                rows = [[row[col] for col in columns] for row in fetched]
                result_sets.append(
                    {
                        "index": idx,
                        "sql": sql,
                        "columns": columns,
                        "rows": rows,
                        "row_count": len(rows),
                        "truncated": truncated,
                    }
                )
            return {
                "script": script,
                "total_result_sets": len(result_sets),
                "results": result_sets,
            }
        except sqlite3.Error as exc:
            raise ServiceError(f"SQL 执行失败: {exc}") from exc
        finally:
            conn.close()

    def export_query_result_csv(
        self,
        columns: List[str],
        rows: List[List[Any]],
        output_file: Path,
    ) -> Path:
        if not columns:
            raise ServiceError("没有可导出的列。")
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with output_file.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for row in rows:
                writer.writerow(row if isinstance(row, list) else list(row))
        return output_file

    def build_statistics(self, questionnaire_id: str) -> Dict[str, Any]:
        _ = questionnaire_id
        raise ServiceError("内置统计已停用，请在“票据与SQL”页通过 SQL 查询生成结果。")

    def export_submissions_csv(self, questionnaire_id: str, output_file: Path) -> Path:
        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            raise ServiceError("问卷不存在。")
        payloads = self.decrypt_submission_payloads(questionnaire_id)
        questions = questionnaire["schema"].get("questions", [])
        repeat_keys_map: Dict[str, List[str]] = {}
        for q in questions:
            qid = str(q.get("id", "")).strip()
            if not qid:
                continue
            if not str(q.get("repeat_from", "")).strip():
                continue
            seen: set[str] = set()
            keys: List[str] = []
            for payload in payloads:
                answers = payload.get("answers", {})
                if not isinstance(answers, dict):
                    continue
                raw_map = answers.get(qid)
                if not isinstance(raw_map, dict):
                    continue
                for raw_key in raw_map.keys():
                    key = str(raw_key).strip()
                    if not key or key in seen:
                        continue
                    seen.add(key)
                    keys.append(key)
            repeat_keys_map[qid] = keys

        headers = [
            "submission_id",
            "questionnaire_id",
            "questionnaire_version",
            "name",
            "code",
            "source",
            "roster_id",
            "member_key",
        ]
        question_headers: List[str] = []
        for q in questions:
            qid = str(q.get("id", "")).strip()
            if not qid:
                continue
            if str(q.get("repeat_from", "")).strip():
                question_headers.extend([f"{qid}::{key}" for key in repeat_keys_map.get(qid, [])])
            else:
                question_headers.append(qid)
        headers.extend(question_headers)

        output_file.parent.mkdir(parents=True, exist_ok=True)
        with output_file.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for payload in payloads:
                verified = payload.get("verified", {})
                row = {
                    "submission_id": payload.get("submission_id", ""),
                    "questionnaire_id": payload.get("questionnaire_id", ""),
                    "questionnaire_version": payload.get("questionnaire_version", ""),
                    "name": payload.get("respondent", {}).get("name", ""),
                    "code": payload.get("respondent", {}).get("code", ""),
                    "source": payload.get("submitted_from", ""),
                    "roster_id": verified.get("roster_id", ""),
                    "member_key": verified.get("member_key", ""),
                }
                answers = payload.get("answers", {})
                if not isinstance(answers, dict):
                    answers = {}
                for q in questions:
                    qid = str(q.get("id", "")).strip()
                    if not qid:
                        continue
                    value = answers.get(qid, "")
                    if str(q.get("repeat_from", "")).strip():
                        value_map = value if isinstance(value, dict) else {}
                        for repeat_key in repeat_keys_map.get(qid, []):
                            cell = value_map.get(repeat_key, "")
                            if isinstance(cell, list):
                                cell = " | ".join(str(v) for v in cell)
                            row[f"{qid}::{repeat_key}"] = cell
                        continue
                    if isinstance(value, list):
                        value = " | ".join(str(v) for v in value)
                    row[qid] = value
                writer.writerow(row)
        return output_file

    def import_vote_file(self, vote_path: Path) -> Tuple[bool, str]:
        if not self.crypto.unlocked:
            raise ServiceError("管理员密钥尚未解锁。")
        if vote_path.suffix.lower() != ".vote":
            return False, f"{vote_path.name}: 不是 .vote 文件"
        try:
            envelope = self.crypto.load_vote_file(vote_path)
            payload = self.crypto.decrypt_envelope(envelope)
        except VoteCryptoError as exc:
            return False, f"{vote_path.name}: 解密失败 ({exc})"

        submission_id = str(payload.get("submission_id", "")).strip()
        questionnaire_id = str(payload.get("questionnaire_id", "")).strip()
        if not submission_id or not questionnaire_id:
            return False, f"{vote_path.name}: 缺少 submission_id 或 questionnaire_id"
        if self.db.submission_exists(submission_id):
            return False, f"{vote_path.name}: 已存在，已跳过"

        questionnaire = self.get_questionnaire(questionnaire_id)
        if not questionnaire:
            return False, f"{vote_path.name}: 对应问卷 {questionnaire_id} 不存在"

        respondent = payload.get("respondent", {})
        context = payload.get("context", {})
        verified = payload.get("verified", {})
        answers_raw = payload.get("answers", {})
        if not isinstance(answers_raw, dict):
            return False, f"{vote_path.name}: 票据 answers 格式错误"

        respondent_name = str(respondent.get("name", "")).strip()
        respondent_code = str(respondent.get("code", "")).strip()
        verified_member_key = str(verified.get("member_key", "")).strip()
        current_member_key = (
            str(context.get("current_member_key", "")).strip()
            or verified_member_key
            or self._resolve_current_member_key(
                questionnaire=questionnaire,
                verified_member_key=verified_member_key,
                respondent_code=respondent_code,
                respondent_name=respondent_name,
                respondent_identity=respondent.get("identity_data", {}),
            )
        )
        answers_for_validate = dict(answers_raw)
        schema = questionnaire.get("schema", {}) if isinstance(questionnaire.get("schema"), dict) else {}
        self._inject_schema_list_sources(
            schema=schema,
            answers_for_validate=answers_for_validate,
            current_member_key=current_member_key,
        )
        self._inject_legacy_repeat_source(schema=schema, answers_for_validate=answers_for_validate)
        ok, rule_errors, cleaned_answers = validate_answers(schema, answers_for_validate)
        if not ok:
            return False, f"{vote_path.name}: 票据内容不符合当前问卷规则（{'；'.join(rule_errors[:4])}）"
        allow_repeat = bool(questionnaire.get("allow_repeat", False))
        allow_same_device_repeat = self._questionnaire_allow_same_device_repeat(questionnaire)
        identity_data_for_dedupe = respondent.get("identity_data", {}) if isinstance(respondent.get("identity_data", {}), dict) else {}
        identity_dedupe_key = self._identity_dedupe_key(identity_data_for_dedupe)
        dedupe_member_key = verified_member_key or current_member_key or identity_dedupe_key
        dedupe_code = respondent_code or str(verified.get("member_code", "")).strip() or identity_dedupe_key
        dedupe_token = str(respondent.get("client_token", "")).strip()
        if not allow_repeat and self.db.detect_duplicate(
            questionnaire_id=questionnaire_id,
            client_token="" if allow_same_device_repeat else dedupe_token,
            respondent_code=dedupe_code,
            verified_member_key=dedupe_member_key,
        ):
            if allow_same_device_repeat:
                return False, f"{vote_path.name}: 当前身份已提交，不能重复归票"
            return False, f"{vote_path.name}: 当前身份或设备已提交，不能重复归票"
        sql_rule_errors = self._evaluate_sql_validation_rules(
            questionnaire=questionnaire,
            answers=cleaned_answers,
            respondent_name=respondent_name,
            respondent_code=respondent_code,
            respondent_identity=respondent.get("identity_data", {}),
            current_member_key=current_member_key,
        )
        if sql_rule_errors:
            return False, f"{vote_path.name}: 联合规则未通过（{'；'.join(sql_rule_errors[:2])}）"

        dst = self._vote_file_path(questionnaire_id, submission_id)
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(vote_path, dst)

        self.db.save_submission_meta(
            submission_id=submission_id,
            questionnaire_id=questionnaire_id,
            questionnaire_version=int(payload.get("questionnaire_version", 1) or 1),
            respondent_name=respondent.get("name", ""),
            respondent_code=respondent.get("code", ""),
            anonymous=False,
            source="offline_import",
            vote_file=str(dst),
            client_token=respondent.get("client_token", ""),
            session_label=context.get("relation_type", ""),
            target_label=context.get("target_label", ""),
            roster_id=verified.get("roster_id", ""),
            verified_member_key=dedupe_member_key,
        )
        self.db.append_audit_log(
            "vote_imported",
            {"submission_id": submission_id, "questionnaire_id": questionnaire_id, "file": str(vote_path)},
        )
        return True, f"{vote_path.name}: 导入成功"

    def create_backup(self, output_file: Path) -> Path:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        if output_file.suffix.lower() != ".zip":
            output_file = output_file.with_suffix(".zip")
        with zipfile.ZipFile(output_file, "w", zipfile.ZIP_DEFLATED) as zf:
            if self.paths.db_file.exists():
                zf.write(self.paths.db_file, "data/votefree.db")
            for folder, arc_root in [
                (self.paths.votes_dir, "data/votes"),
                (self.paths.keys_dir, "data/keys"),
                (self.paths.exports_dir, "data/exports"),
            ]:
                if not folder.exists():
                    continue
                for file_path in folder.rglob("*"):
                    if file_path.is_file():
                        zf.write(file_path, f"{arc_root}/{file_path.relative_to(folder).as_posix()}")
        self.db.append_audit_log("backup_created", {"file": str(output_file)})
        return output_file

    def summary_cards(self) -> Dict[str, Any]:
        return {
            "questionnaires": self.db.count_questionnaires(),
            "submissions": self.db.count_submissions(),
            "rosters": self.db.count_rosters(),
            "votes_dir": str(self.paths.votes_dir),
        }
